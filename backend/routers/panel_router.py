"""
Panel Router - Custom Panel System (Advanced Business Tools)
Handles panel CRUD and field management.
Only available to users with businessToolAccess = "advanced".
Panel creation restricted to seller admins (not employees).
"""

from fastapi import APIRouter, HTTPException, Header
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone
from bson import ObjectId
from pydantic import BaseModel, Field
import logging
import re
import io

logger = logging.getLogger(__name__)

# Limits
MAX_PANELS_PER_BUSINESS = 10
MAX_FIELDS_PER_PANEL = 20

MAX_RECORDS_PER_PAGE = 50

VALID_FIELD_TYPES = {"text", "number", "date", "dropdown", "multiselect", "boolean", "longtext", "relation"}
VALID_RELATION_TYPES = {"many_to_one", "one_to_one"}
SYSTEM_LINKABLE = {"inventory", "invoices", "buyers", "suppliers", "purchase_orders", "quotations", "composite_products", "employees"}


# ── Pydantic Models ──

class PanelFieldInput(BaseModel):
    key: str = Field(..., min_length=1, max_length=50)
    label: str = Field(..., min_length=1, max_length=100)
    type: str = Field(..., description="text, number, date, dropdown, multiselect, boolean, longtext, relation")
    required: bool = False
    unique: bool = False
    options: Optional[List[str]] = None
    relatedPanel: Optional[str] = None
    relationType: Optional[str] = None
    bindingField: Optional[str] = None
    order: Optional[int] = 0


class CreatePanelRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    description: Optional[str] = ""
    icon: Optional[str] = "layout-grid"
    color: Optional[str] = "blue"
    fields: Optional[List[PanelFieldInput]] = []
    allowedModules: Optional[List[str]] = Field(default_factory=list)
    allowedPanels: Optional[List[str]] = Field(default_factory=list)


class UpdatePanelRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    icon: Optional[str] = None
    color: Optional[str] = None
    allowedModules: Optional[List[str]] = None
    allowedPanels: Optional[List[str]] = None


class AddFieldRequest(BaseModel):
    key: str = Field(..., min_length=1, max_length=50)
    label: str = Field(..., min_length=1, max_length=100)
    type: str
    required: bool = False
    unique: bool = False
    options: Optional[List[str]] = None
    relatedPanel: Optional[str] = None
    relationType: Optional[str] = None
    bindingField: Optional[str] = None


class UpdateFieldRequest(BaseModel):
    label: Optional[str] = None
    required: Optional[bool] = None
    options: Optional[List[str]] = None
    disabled: Optional[bool] = None


class ReorderFieldsRequest(BaseModel):
    fieldKeys: List[str]


class CreateRecordRequest(BaseModel):
    data: dict


class UpdateRecordRequest(BaseModel):
    data: dict


def init_panel_router(db, verify_token_func, automation_executor=None):
    from utils.permissions import authenticate_user, resolve_seller_id, is_platform_admin, normalize_permissions

    router = APIRouter(tags=["Panels"])

    def serialize_doc(doc):
        if doc is None:
            return None
        if isinstance(doc, list):
            return [serialize_doc(d) for d in doc]
        if isinstance(doc, dict):
            result = {}
            for key, value in doc.items():
                if key == "_id":
                    result["id"] = str(value)
                elif isinstance(value, ObjectId):
                    result[key] = str(value)
                elif isinstance(value, datetime):
                    result[key] = value.isoformat()
                elif isinstance(value, dict):
                    result[key] = serialize_doc(value)
                elif isinstance(value, list):
                    result[key] = serialize_doc(value)
                else:
                    result[key] = value
            return result
        return doc

    async def get_current_user(authorization: str = Header(...)):
        return await authenticate_user(db, verify_token_func, authorization)

    async def get_seller_id(user: dict) -> str:
        sid = resolve_seller_id(user)
        if sid is None:
            # Platform admin: use their own user ID for panel ownership
            uid = user.get("_id") or user.get("id")
            return str(uid)
        return sid

    def require_advanced_access(user: dict):
        """Check that the user has advanced business tool access."""
        access_level = user.get("businessToolAccess", "standard")
        if access_level != "advanced" and not is_platform_admin(user):
            raise HTTPException(status_code=403, detail="Advanced access required. Contact platform admin to upgrade.")

    def require_seller_admin(user: dict):
        """Only seller admin (not employee) can create/modify panel structure."""
        if user.get("companyId") and user.get("employeeStatus") == "active":
            raise HTTPException(status_code=403, detail="Only the business admin can manage panel structure.")
        if not is_platform_admin(user) and user.get("accountType") not in (None, "seller"):
            raise HTTPException(status_code=403, detail="Only the business admin can manage panel structure.")

    def make_slug(name: str) -> str:
        slug = re.sub(r'[^a-z0-9]+', '-', name.lower().strip()).strip('-')
        return slug or "panel"

    def validate_field(field: dict):
        if field["type"] not in VALID_FIELD_TYPES:
            raise HTTPException(status_code=400, detail=f"Invalid field type: {field['type']}. Must be one of {VALID_FIELD_TYPES}")
        if field["type"] in ("dropdown", "multiselect") and not field.get("options"):
            raise HTTPException(status_code=400, detail=f"Field '{field['key']}' requires options for {field['type']} type")
        if field["type"] == "relation":
            if not field.get("relatedPanel"):
                raise HTTPException(status_code=400, detail=f"Field '{field['key']}' requires relatedPanel for relation type")
            rt = field.get("relationType", "many_to_one")
            if rt not in VALID_RELATION_TYPES:
                raise HTTPException(status_code=400, detail=f"Invalid relationType: {rt}")

    async def validate_allowed_panels(seller_id: str, allowed_panels: list, current_panel_id: str = None):
        """Validate panel linking rules: no self-link, no circular, max 2, must exist."""
        if not allowed_panels:
            return
        if len(allowed_panels) > 2:
            raise HTTPException(status_code=400, detail="Maximum 2 linked panels allowed.")
        if len(set(allowed_panels)) != len(allowed_panels):
            raise HTTPException(status_code=400, detail="Duplicate panel links are not allowed.")
        if current_panel_id and current_panel_id in allowed_panels:
            raise HTTPException(status_code=400, detail="A panel cannot link to itself.")
        for pid in allowed_panels:
            try:
                target = await db.panels.find_one(
                    {"_id": ObjectId(pid), "sellerId": ObjectId(seller_id)},
                    {"_id": 1, "allowedPanels": 1}
                )
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid panel ID: {pid}")
            if not target:
                raise HTTPException(status_code=404, detail=f"Linked panel not found: {pid}")
            # Circular check: if the target already links back to us
            if current_panel_id:
                target_links = target.get("allowedPanels", [])
                if current_panel_id in target_links:
                    raise HTTPException(status_code=400, detail=f"Circular linking detected: panel '{pid}' already links back to this panel.")

    # ── LIST PANELS ──
    @router.get("/panels")
    async def list_panels(authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        seller_id = await get_seller_id(user)

        cursor = db.panels.find(
            {"sellerId": ObjectId(seller_id)},
            {"_id": 1, "name": 1, "slug": 1, "description": 1, "icon": 1, "color": 1,
             "fields": 1, "allowedModules": 1, "allowedPanels": 1, "createdAt": 1, "updatedAt": 1}
        ).sort("createdAt", 1)
        panels = await cursor.to_list(MAX_PANELS_PER_BUSINESS + 5)
        return {"panels": serialize_doc(panels), "count": len(panels), "limit": MAX_PANELS_PER_BUSINESS}

    # ── LINKABLE TARGETS (for relation fields) — must be before {panel_id} routes ──
    @router.get("/panels/linkable-targets")
    async def get_linkable_targets(authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        seller_id = await get_seller_id(user)

        targets = [
            {"id": "inventory", "name": "Inventory", "type": "system"},
            {"id": "invoices", "name": "Invoices", "type": "system"},
            {"id": "buyers", "name": "Buyers", "type": "system"},
            {"id": "suppliers", "name": "Suppliers", "type": "system"},
            {"id": "purchase_orders", "name": "Purchase Orders", "type": "system"},
            {"id": "quotations", "name": "Quotations", "type": "system"},
            {"id": "composite_products", "name": "Composite Products", "type": "system"},
            {"id": "employees", "name": "Employees", "type": "system"},
        ]

        panels = await db.panels.find(
            {"sellerId": ObjectId(seller_id)},
            {"_id": 1, "name": 1}
        ).to_list(MAX_PANELS_PER_BUSINESS)

        for p in panels:
            targets.append({"id": str(p["_id"]), "name": p["name"], "type": "panel"})

        return {"targets": targets}

    # ── MODULE FIELDS (for binding variable selection) ──
    SYSTEM_MODULE_FIELDS = {
        "inventory": [
            {"key": "productName", "label": "Product Name", "type": "text"},
            {"key": "sku", "label": "SKU", "type": "text"},
            {"key": "category", "label": "Category", "type": "text"},
            {"key": "stock", "label": "Stock", "type": "number"},
            {"key": "quantity", "label": "Quantity", "type": "number"},
            {"key": "minStock", "label": "Min Stock", "type": "number"},
            {"key": "reorderPoint", "label": "Reorder Point", "type": "number"},
        ],
        "invoices": [
            {"key": "invoiceNumber", "label": "Invoice Number", "type": "text"},
            {"key": "buyerName", "label": "Buyer Name", "type": "text"},
            {"key": "totalAmount", "label": "Total Amount", "type": "number"},
        ],
        "buyers": [
            {"key": "name", "label": "Buyer Name", "type": "text"},
            {"key": "phone", "label": "Phone", "type": "text"},
            {"key": "email", "label": "Email", "type": "text"},
        ],
        "suppliers": [
            {"key": "name", "label": "Supplier Name", "type": "text"},
            {"key": "phone", "label": "Phone", "type": "text"},
            {"key": "email", "label": "Email", "type": "text"},
        ],
        "purchase_orders": [
            {"key": "poNumber", "label": "PO Number", "type": "text"},
            {"key": "supplierName", "label": "Supplier Name", "type": "text"},
            {"key": "totalAmount", "label": "Total Amount", "type": "number"},
        ],
        "quotations": [
            {"key": "quotationNumber", "label": "Quotation Number", "type": "text"},
            {"key": "buyerName", "label": "Buyer Name", "type": "text"},
            {"key": "totalAmount", "label": "Total Amount", "type": "number"},
        ],
        "composite_products": [
            {"key": "name", "label": "Product Name", "type": "text"},
            {"key": "sku", "label": "SKU", "type": "text"},
        ],
        "employees": [
            {"key": "name", "label": "Employee Name", "type": "text"},
            {"key": "role", "label": "Role", "type": "text"},
            {"key": "email", "label": "Email", "type": "text"},
        ],
    }

    @router.get("/panels/module-fields/{module_id}")
    async def get_module_fields(module_id: str, authorization: str = Header(...)):
        """Return available fields for a system module or custom panel — used for binding variable selection."""
        user = await get_current_user(authorization)
        require_advanced_access(user)
        seller_id = await get_seller_id(user)

        # System module
        if module_id in SYSTEM_MODULE_FIELDS:
            return {"fields": SYSTEM_MODULE_FIELDS[module_id], "type": "system", "name": module_id}

        # Custom panel
        try:
            panel = await db.panels.find_one({"_id": ObjectId(module_id), "sellerId": ObjectId(seller_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid module ID")
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        fields = [
            {"key": f["key"], "label": f["label"], "type": f["type"]}
            for f in panel.get("fields", [])
            if f["type"] != "relation"
        ]
        return {"fields": fields, "type": "panel", "name": panel["name"]}


    # ── RELATED RECORDS (for panel data integration with modules) ──
    @router.get("/panels/related-records")
    async def get_related_records(
        module: str,
        entityId: str,
        authorization: str = Header(...)
    ):
        """Find panel records related to a specific entity (e.g. inventory product)."""
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        # Find panels with relation fields pointing to this module
        panels_with_relations = await db.panels.find({
            "sellerId": ObjectId(seller_id),
            "fields.type": "relation",
            "fields.relatedPanel": module,
        }).to_list(20)

        result = []
        for panel in panels_with_relations:
            panel_id = str(panel["_id"])

            # Check employee permission for this panel
            if user.get("companyId") and user.get("employeeStatus") == "active":
                perms = normalize_permissions(user.get("employeePermissions", {}))
                panel_perms = perms.get("panels", {}).get(panel_id, {})
                if not panel_perms.get("canView"):
                    continue

            # Find relation field keys for this module
            relation_keys = [
                f["key"] for f in panel.get("fields", [])
                if f["type"] == "relation" and f.get("relatedPanel") == module
            ]
            if not relation_keys:
                continue

            # Find records matching the entityId
            query = {
                "panelId": panel["_id"],
                "sellerId": ObjectId(seller_id),
                "$or": [{f"data.{key}": entityId} for key in relation_keys]
            }
            records = await db.panel_records.find(query).limit(20).to_list(20)

            if records:
                resolved_records = []
                for rec in records:
                    display = {}
                    for f in panel.get("fields", []):
                        val = rec.get("data", {}).get(f["key"])
                        if val is not None and f["type"] != "relation":
                            display[f.get("label", f["key"])] = val
                    resolved_records.append({
                        "id": str(rec["_id"]),
                        "data": display,
                    })

                result.append({
                    "panelId": panel_id,
                    "panelName": panel.get("name", "Panel"),
                    "panelColor": panel.get("color", "blue"),
                    "records": resolved_records,
                })

        return {"groups": result}

    # ── GET SINGLE PANEL ──
    @router.get("/panels/{panel_id}")
    async def get_panel(panel_id: str, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        try:
            panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid panel ID")

        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")
        return serialize_doc(panel)

    # ── FIELD VISIBILITY (aggregated from automation rules) ──
    @router.get("/panels/{panel_id}/field-visibility")
    async def get_field_visibility(panel_id: str, authorization: str = Header(...)):
        """Return merged field_visibility settings from all automation rules targeting this panel.
        Most restrictive wins: visible=false wins over visible=true, editable=false wins over editable=true."""
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        rules = await db.automation_rules.find({
            "sellerId": ObjectId(seller_id),
            "is_active": True,
            "targets.target_panel_id": panel_id,
        }).to_list(100)

        merged: dict = {}  # field_key -> {visible, editable, source_rules}
        for rule in rules:
            for t in rule.get("targets", []):
                if t.get("target_panel_id") != panel_id:
                    continue
                for fv in (t.get("field_visibility") or []):
                    fk = fv.get("field", "")
                    if not fk:
                        continue
                    if fk not in merged:
                        merged[fk] = {"visible": fv.get("visible", True), "editable": fv.get("editable", True), "source_rules": []}
                    else:
                        if not fv.get("visible", True):
                            merged[fk]["visible"] = False
                        if not fv.get("editable", True):
                            merged[fk]["editable"] = False
                    merged[fk]["source_rules"].append(rule.get("name", "Unknown"))

        return {"panel_id": panel_id, "field_visibility": merged}

    # ── CREATE PANEL ──
    @router.post("/panels")
    async def create_panel(data: CreatePanelRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        # Check limit
        count = await db.panels.count_documents({"sellerId": ObjectId(seller_id)})
        if count >= MAX_PANELS_PER_BUSINESS:
            raise HTTPException(status_code=400, detail=f"Maximum {MAX_PANELS_PER_BUSINESS} panels allowed per business.")

        # Check duplicate name
        existing = await db.panels.find_one({"sellerId": ObjectId(seller_id), "name": {"$regex": f"^{re.escape(data.name.strip())}$", "$options": "i"}})
        if existing:
            raise HTTPException(status_code=400, detail="A panel with this name already exists.")

        # Validate fields
        fields = []
        seen_keys = set()
        for i, f in enumerate(data.fields or []):
            if len(fields) >= MAX_FIELDS_PER_PANEL:
                raise HTTPException(status_code=400, detail=f"Maximum {MAX_FIELDS_PER_PANEL} fields allowed per panel.")
            fd = f.model_dump()
            validate_field(fd)
            if fd["key"] in seen_keys:
                raise HTTPException(status_code=400, detail=f"Duplicate field key: {fd['key']}")
            seen_keys.add(fd["key"])
            fd["order"] = i
            fields.append(fd)

        now = datetime.now(timezone.utc)
        slug = make_slug(data.name)

        # Ensure slug uniqueness
        slug_exists = await db.panels.find_one({"sellerId": ObjectId(seller_id), "slug": slug})
        if slug_exists:
            slug = f"{slug}-{int(now.timestamp()) % 10000}"

        doc = {
            "sellerId": ObjectId(seller_id),
            "name": data.name.strip(),
            "slug": slug,
            "description": (data.description or "").strip(),
            "icon": data.icon or "layout-grid",
            "color": data.color or "blue",
            "fields": fields,
            "allowedModules": data.allowedModules or [],
            "allowedPanels": data.allowedPanels or [],
            "createdAt": now,
            "updatedAt": now,
        }

        # Validate allowedPanels
        if data.allowedPanels:
            await validate_allowed_panels(seller_id, data.allowedPanels, current_panel_id=None)

        # Auto-add required inventory relation field if linked to inventory
        if "inventory" in (data.allowedModules or []):
            has_inv_relation = any(
                f["type"] == "relation" and f.get("relatedPanel") == "inventory"
                for f in fields
            )
            if not has_inv_relation:
                inv_field = {
                    "key": "product",
                    "label": "Product (Linked to Inventory)",
                    "type": "relation",
                    "required": True,
                    "unique": False,
                    "relatedPanel": "inventory",
                    "relationType": "many_to_one",
                    "bindingField": "productName",
                    "options": None,
                    "order": 0,
                    "systemManaged": True,
                }
                # Shift existing field orders
                for f in doc["fields"]:
                    f["order"] = f.get("order", 0) + 1
                doc["fields"].insert(0, inv_field)

        # Auto-add required invoice relation field if linked to invoices
        if "invoices" in (data.allowedModules or []):
            has_inv_relation = any(
                f["type"] == "relation" and f.get("relatedPanel") == "invoices"
                for f in doc["fields"]
            )
            if not has_inv_relation:
                inv_field = {
                    "key": "invoice",
                    "label": "Invoice (Linked to Invoices)",
                    "type": "relation",
                    "required": True,
                    "unique": False,
                    "relatedPanel": "invoices",
                    "relationType": "many_to_one",
                    "bindingField": "invoiceNumber",
                    "options": None,
                    "order": len(doc["fields"]),
                    "systemManaged": True,
                }
                doc["fields"].append(inv_field)

        result = await db.panels.insert_one(doc)
        doc["_id"] = result.inserted_id
        logger.info(f"Panel created: {data.name} for seller {seller_id}")
        return serialize_doc(doc)

    # ── UPDATE PANEL (name, description, icon, color) ──
    @router.put("/panels/{panel_id}")
    async def update_panel(panel_id: str, data: UpdatePanelRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        try:
            panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid panel ID")
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        update = {"updatedAt": datetime.now(timezone.utc)}
        if data.name is not None:
            # Check duplicate
            existing = await db.panels.find_one({
                "sellerId": ObjectId(seller_id),
                "name": {"$regex": f"^{re.escape(data.name.strip())}$", "$options": "i"},
                "_id": {"$ne": ObjectId(panel_id)}
            })
            if existing:
                raise HTTPException(status_code=400, detail="A panel with this name already exists.")
            update["name"] = data.name.strip()
            update["slug"] = make_slug(data.name)
        if data.description is not None:
            update["description"] = data.description.strip()
        if data.icon is not None:
            update["icon"] = data.icon
        if data.color is not None:
            update["color"] = data.color
        if data.allowedModules is not None:
            update["allowedModules"] = data.allowedModules
            current_fields = panel.get("fields", [])
            # Auto-add inventory relation field if newly linking to inventory
            if "inventory" in data.allowedModules:
                has_inv_relation = any(
                    f["type"] == "relation" and f.get("relatedPanel") == "inventory"
                    for f in current_fields
                )
                if not has_inv_relation:
                    inv_field = {
                        "key": "product", "label": "Product (Linked to Inventory)", "type": "relation",
                        "required": True, "unique": False,
                        "relatedPanel": "inventory", "relationType": "many_to_one",
                        "bindingField": "productName",
                        "options": None, "order": 0, "systemManaged": True,
                    }
                    for f in current_fields:
                        f["order"] = f.get("order", 0) + 1
                    current_fields.insert(0, inv_field)
                    update["fields"] = current_fields

            # Auto-add invoice relation field if newly linking to invoices
            if "invoices" in data.allowedModules:
                fields_to_check = update.get("fields", current_fields)
                has_inv_relation = any(
                    f["type"] == "relation" and f.get("relatedPanel") == "invoices"
                    for f in fields_to_check
                )
                if not has_inv_relation:
                    inv_field = {
                        "key": "invoice", "label": "Invoice (Linked to Invoices)", "type": "relation",
                        "required": True, "unique": False,
                        "relatedPanel": "invoices", "relationType": "many_to_one",
                        "bindingField": "invoiceNumber",
                        "options": None, "order": len(fields_to_check), "systemManaged": True,
                    }
                    if "fields" in update:
                        update["fields"].append(inv_field)
                    else:
                        current_fields.append(inv_field)
                        update["fields"] = current_fields

        if data.allowedPanels is not None:
            await validate_allowed_panels(seller_id, data.allowedPanels, current_panel_id=panel_id)
            update["allowedPanels"] = data.allowedPanels

        await db.panels.update_one({"_id": ObjectId(panel_id)}, {"$set": update})
        logger.info(f"Panel {panel_id} updated")
        return {"message": "Panel updated"}

    # ── DELETE PANEL ──
    @router.delete("/panels/{panel_id}")
    async def delete_panel(panel_id: str, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        try:
            panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid panel ID")
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        # Check for linked records
        record_count = await db.panel_records.count_documents({"panelId": ObjectId(panel_id)})
        if record_count > 0:
            raise HTTPException(status_code=400, detail=f"Cannot delete panel with {record_count} existing records. Delete records first.")

        # Check if other panels have relation fields pointing to this panel
        referencing = await db.panels.find_one({
            "sellerId": ObjectId(seller_id),
            "fields.relatedPanel": panel_id,
            "_id": {"$ne": ObjectId(panel_id)}
        })
        if referencing:
            raise HTTPException(status_code=400, detail=f"Cannot delete: panel '{referencing['name']}' has a relation field linked to this panel.")

        await db.panels.delete_one({"_id": ObjectId(panel_id)})
        logger.info(f"Panel {panel_id} deleted")
        return {"message": "Panel deleted"}

    # ── ADD FIELD TO PANEL ──
    @router.post("/panels/{panel_id}/fields")
    async def add_field(panel_id: str, data: AddFieldRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        fields = panel.get("fields", [])
        if len(fields) >= MAX_FIELDS_PER_PANEL:
            raise HTTPException(status_code=400, detail=f"Maximum {MAX_FIELDS_PER_PANEL} fields allowed per panel.")

        # Check duplicate key
        if any(f["key"] == data.key for f in fields):
            raise HTTPException(status_code=400, detail=f"Field key '{data.key}' already exists in this panel.")

        fd = data.model_dump()
        validate_field(fd)
        fd["order"] = len(fields)

        await db.panels.update_one(
            {"_id": ObjectId(panel_id)},
            {"$push": {"fields": fd}, "$set": {"updatedAt": datetime.now(timezone.utc)}}
        )
        logger.info(f"Field '{data.key}' added to panel {panel_id}")
        return {"message": "Field added", "field": fd}

    # ── UPDATE FIELD ──
    @router.put("/panels/{panel_id}/fields/{field_key}")
    async def update_field(panel_id: str, field_key: str, data: UpdateFieldRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        fields = panel.get("fields", [])
        field_idx = next((i for i, f in enumerate(fields) if f["key"] == field_key), None)
        if field_idx is None:
            raise HTTPException(status_code=404, detail=f"Field '{field_key}' not found")

        update_ops = {"updatedAt": datetime.now(timezone.utc)}
        if data.label is not None:
            update_ops[f"fields.{field_idx}.label"] = data.label
        if data.required is not None:
            update_ops[f"fields.{field_idx}.required"] = data.required
        if data.options is not None:
            update_ops[f"fields.{field_idx}.options"] = data.options
        if data.disabled is not None:
            update_ops[f"fields.{field_idx}.disabled"] = data.disabled

        await db.panels.update_one({"_id": ObjectId(panel_id)}, {"$set": update_ops})
        return {"message": "Field updated"}

    # ── DELETE FIELD ──
    @router.delete("/panels/{panel_id}/fields/{field_key}")
    async def delete_field(panel_id: str, field_key: str, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        fields = panel.get("fields", [])
        if not any(f["key"] == field_key for f in fields):
            raise HTTPException(status_code=404, detail=f"Field '{field_key}' not found")

        # Block deletion if any record has data for this field
        has_data = await db.panel_records.find_one({
            "panelId": ObjectId(panel_id),
            "sellerId": ObjectId(seller_id),
            f"data.{field_key}": {"$exists": True, "$nin": [None, ""]}
        })
        if has_data:
            raise HTTPException(status_code=400, detail=f"Cannot delete field '{field_key}': records contain data for this field. Disable it instead.")

        new_fields = [f for f in fields if f["key"] != field_key]
        # Re-order
        for i, f in enumerate(new_fields):
            f["order"] = i

        await db.panels.update_one(
            {"_id": ObjectId(panel_id)},
            {"$set": {"fields": new_fields, "updatedAt": datetime.now(timezone.utc)}}
        )
        return {"message": "Field deleted"}

    # ── REORDER FIELDS ──
    @router.put("/panels/{panel_id}/fields-order")
    async def reorder_fields(panel_id: str, data: ReorderFieldsRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        require_advanced_access(user)
        require_seller_admin(user)
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        fields = panel.get("fields", [])
        field_map = {f["key"]: f for f in fields}
        reordered = []
        for i, key in enumerate(data.fieldKeys):
            if key not in field_map:
                raise HTTPException(status_code=400, detail=f"Field key '{key}' not found")
            f = field_map[key]
            f["order"] = i
            reordered.append(f)

        # Add any fields not in the reorder list at the end
        for f in fields:
            if f["key"] not in data.fieldKeys:
                f["order"] = len(reordered)
                reordered.append(f)

        await db.panels.update_one(
            {"_id": ObjectId(panel_id)},
            {"$set": {"fields": reordered, "updatedAt": datetime.now(timezone.utc)}}
        )
        return {"message": "Fields reordered"}

    # ═══════════════════════════════════════════
    # RECORD CRUD — Phase 2
    # ═══════════════════════════════════════════

    def check_panel_access(user: dict, panel_id: str, action: str = "view"):
        """Unified panel access check. Admin/seller with advanced access and employees with panel permissions."""
        if user.get("accountType") == "buyer":
            raise HTTPException(status_code=403, detail="Buyers cannot access panels.")
        if is_platform_admin(user):
            return
        # Employee with panel permissions
        if user.get("companyId") and user.get("employeeStatus") == "active":
            perms = normalize_permissions(user.get("employeePermissions", {}))
            panel_perms = perms.get("panels", {}).get(panel_id, {})
            perm_map = {"view": "canView", "create": "canCreate", "edit": "canEdit"}
            perm_key = perm_map.get(action, "canView")
            if not panel_perms.get(perm_key):
                raise HTTPException(status_code=403, detail=f"You don't have {action} permission for this panel")
            return
        # Seller admin - require advanced access
        require_advanced_access(user)

    async def resolve_relation_display(seller_id: str, field: dict, value):
        """Resolve a relation value to a display label."""
        if not value:
            return None
        target = field.get("relatedPanel", "")
        try:
            if target == "inventory":
                listing = await db.sellerListings.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"productId": 1, "sku": 1})
                if listing:
                    product = await db.products.find_one({"_id": listing.get("productId")}, {"name": 1}) if listing.get("productId") else None
                    return {"id": str(listing["_id"]), "label": product.get("name", "") if product else "", "sku": listing.get("sku", "")}
            elif target == "invoices":
                doc = await db.invoices.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"invoiceNumber": 1, "buyerName": 1})
                if doc:
                    return {"id": str(doc["_id"]), "label": doc.get("invoiceNumber", ""), "sub": doc.get("buyerName", "")}
            elif target == "buyers":
                doc = await db.seller_buyers.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"buyerName": 1, "phone": 1})
                if doc:
                    return {"id": str(doc["_id"]), "label": doc.get("buyerName", ""), "sub": doc.get("phone", "")}
            elif target == "suppliers":
                doc = await db.seller_suppliers.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"supplierName": 1, "phone": 1})
                if doc:
                    return {"id": str(doc["_id"]), "label": doc.get("supplierName", ""), "sub": doc.get("phone", "")}
            elif target == "purchase_orders":
                doc = await db.purchase_orders.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"poNumber": 1, "supplierName": 1})
                if doc:
                    return {"id": str(doc["_id"]), "label": doc.get("poNumber", ""), "sub": doc.get("supplierName", "")}
            elif target == "quotations":
                doc = await db.quotations.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"quotationNumber": 1, "buyerId": 1})
                if doc:
                    buyer_name = ""
                    if doc.get("buyerId"):
                        buyer = await db.seller_buyers.find_one({"_id": doc["buyerId"]}, {"buyerName": 1})
                        buyer_name = buyer.get("buyerName", "") if buyer else ""
                    return {"id": str(doc["_id"]), "label": doc.get("quotationNumber", ""), "sub": buyer_name}
            elif target == "composite_products":
                doc = await db.composite_products.find_one({"_id": ObjectId(value), "sellerId": ObjectId(seller_id)}, {"name": 1, "price": 1})
                if doc:
                    return {"id": str(doc["_id"]), "label": doc.get("name", ""), "sub": f"Rs {doc['price']}" if doc.get("price") else ""}
            elif target == "employees":
                doc = await db.users.find_one({"_id": ObjectId(value), "companyId": ObjectId(seller_id), "employeeStatus": {"$in": ["active", "disabled"]}}, {"name": 1, "email": 1, "employeeRole": 1})
                if doc:
                    return {"id": str(doc["_id"]), "label": doc.get("name", doc.get("email", "")), "sub": doc.get("employeeRole", "")}
            else:
                # Custom panel
                doc = await db.panel_records.find_one({"_id": ObjectId(value), "panelId": ObjectId(target), "sellerId": ObjectId(seller_id)})
                if doc:
                    data = doc.get("data", {})
                    linked_panel = await db.panels.find_one({"_id": ObjectId(target)}, {"fields": 1, "name": 1})
                    label = ""
                    if linked_panel:
                        for f in linked_panel.get("fields", []):
                            if f["type"] in ("text", "dropdown") and data.get(f["key"]):
                                label = str(data[f["key"]])
                                break
                    return {"id": str(doc["_id"]), "label": label or str(doc["_id"])[:8]}
        except Exception:
            pass
        return {"id": str(value), "label": str(value)[:12]}

    async def validate_record_data(panel: dict, data: dict, seller_id: str, exclude_record_id: str = None):
        """Validate record data against panel field definitions."""
        fields = panel.get("fields", [])
        errors = []

        for f in fields:
            key = f["key"]
            val = data.get(key)
            is_disabled = f.get("disabled", False)
            if is_disabled:
                continue

            if f.get("required") and (val is None or val == "" or val == []):
                errors.append(f"Field '{f['label']}' is required")
                continue

            if val is None or val == "":
                continue

            ftype = f["type"]
            if ftype == "number":
                try:
                    float(val)
                except (ValueError, TypeError):
                    errors.append(f"Field '{f['label']}' must be a number")

            elif ftype == "dropdown":
                opts = f.get("options", [])
                if opts and str(val) not in opts:
                    errors.append(f"Field '{f['label']}' must be one of: {', '.join(opts)}")

            elif ftype == "multiselect":
                if not isinstance(val, list):
                    errors.append(f"Field '{f['label']}' must be a list")
                else:
                    opts = f.get("options", [])
                    if opts:
                        invalid = [v for v in val if v not in opts]
                        if invalid:
                            errors.append(f"Field '{f['label']}' has invalid options: {', '.join(invalid)}")

            elif ftype == "boolean":
                if not isinstance(val, bool):
                    errors.append(f"Field '{f['label']}' must be true or false")

            elif ftype == "relation":
                target = f.get("relatedPanel", "")
                if target and val:
                    exists = False
                    try:
                        if target == "inventory":
                            exists = bool(await db.sellerListings.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id), "status": {"$in": ["active", "paused"]}}, {"_id": 1}))
                        elif target == "invoices":
                            exists = bool(await db.invoices.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                        elif target == "buyers":
                            exists = bool(await db.seller_buyers.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                        elif target == "suppliers":
                            exists = bool(await db.seller_suppliers.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                        elif target == "purchase_orders":
                            exists = bool(await db.purchase_orders.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                        elif target == "quotations":
                            exists = bool(await db.quotations.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                        elif target == "composite_products":
                            exists = bool(await db.composite_products.find_one({"_id": ObjectId(val), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                        elif target == "employees":
                            exists = bool(await db.users.find_one({"_id": ObjectId(val), "companyId": ObjectId(seller_id), "employeeStatus": {"$in": ["active", "disabled"]}}, {"_id": 1}))
                        else:
                            exists = bool(await db.panel_records.find_one({"_id": ObjectId(val), "panelId": ObjectId(target), "sellerId": ObjectId(seller_id)}, {"_id": 1}))
                    except Exception:
                        pass
                    if not exists:
                        errors.append(f"Field '{f['label']}' references a non-existent record")

                    if f.get("relationType") == "one_to_one" and val:
                        dup_query = {
                            "panelId": ObjectId(str(panel["_id"])),
                            "sellerId": ObjectId(seller_id),
                            f"data.{key}": val
                        }
                        if exclude_record_id:
                            dup_query["_id"] = {"$ne": ObjectId(exclude_record_id)}
                        existing = await db.panel_records.find_one(dup_query, {"_id": 1})
                        if existing:
                            errors.append(f"Field '{f['label']}' already has a one-to-one link to this record")

            # Unique field validation
            if f.get("unique") and val is not None and val != "":
                dup_query = {
                    "panelId": ObjectId(str(panel["_id"])),
                    "sellerId": ObjectId(seller_id),
                    f"data.{key}": val,
                }
                if exclude_record_id:
                    dup_query["_id"] = {"$ne": ObjectId(exclude_record_id)}
                dup = await db.panel_records.find_one(dup_query, {"_id": 1})
                if dup:
                    errors.append(f"Field '{f['label']}' value '{val}' already exists. Must be unique.")

        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))

    # ── LIST RECORDS ──
    @router.get("/panels/{panel_id}/records")
    async def list_records(
        panel_id: str,
        page: int = 1,
        search: str = "",
        authorization: str = Header(...)
    ):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        query = {"panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)}
        if search:
            # Search across text fields
            text_keys = [f"data.{f['key']}" for f in panel.get("fields", []) if f["type"] in ("text", "longtext", "dropdown")]
            if text_keys:
                query["$or"] = [{k: {"$regex": search, "$options": "i"}} for k in text_keys]

        total = await db.panel_records.count_documents(query)
        skip = (max(1, page) - 1) * MAX_RECORDS_PER_PAGE

        cursor = db.panel_records.find(query).sort("createdAt", -1).skip(skip).limit(MAX_RECORDS_PER_PAGE)
        records = await cursor.to_list(MAX_RECORDS_PER_PAGE)

        # Resolve relation fields for display
        relation_fields = [f for f in panel.get("fields", []) if f["type"] == "relation"]
        serialized = []
        for rec in records:
            sr = serialize_doc(rec)
            sr["_resolved"] = {}
            for rf in relation_fields:
                val = rec.get("data", {}).get(rf["key"])
                if val:
                    sr["_resolved"][rf["key"]] = await resolve_relation_display(seller_id, rf, val)
            serialized.append(sr)

        return {
            "records": serialized,
            "total": total,
            "page": page,
            "pages": max(1, (total + MAX_RECORDS_PER_PAGE - 1) // MAX_RECORDS_PER_PAGE),
            "panelName": panel["name"],
        }

    # ── GET SINGLE RECORD ──
    @router.get("/panels/{panel_id}/records/{record_id}")
    async def get_record(panel_id: str, record_id: str, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        record = await db.panel_records.find_one({"_id": ObjectId(record_id), "panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")

        sr = serialize_doc(record)
        sr["_resolved"] = {}
        for rf in panel.get("fields", []):
            if rf["type"] == "relation":
                val = record.get("data", {}).get(rf["key"])
                if val:
                    sr["_resolved"][rf["key"]] = await resolve_relation_display(seller_id, rf, val)

        return {"record": sr, "panel": serialize_doc(panel)}

    # ── CREATE RECORD ──
    @router.post("/panels/{panel_id}/records")
    async def create_record(panel_id: str, data: CreateRecordRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "create")
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        await validate_record_data(panel, data.data, seller_id, exclude_record_id=None)

        field_keys = {f["key"] for f in panel.get("fields", []) if not f.get("disabled")}
        clean_data = {k: v for k, v in data.data.items() if k in field_keys}

        now = datetime.now(timezone.utc)
        user_id = str(user.get("_id") or user.get("id"))

        doc = {
            "panelId": ObjectId(panel_id),
            "sellerId": ObjectId(seller_id),
            "data": clean_data,
            "createdBy": user_id,
            "createdAt": now,
            "updatedAt": now,
        }

        # Extract entity_id from relation fields for linking
        for f in panel.get("fields", []):
            if f.get("type") == "relation" and clean_data.get(f["key"]):
                doc["entity_id"] = clean_data[f["key"]]
                break

        result = await db.panel_records.insert_one(doc)
        doc["_id"] = result.inserted_id

        # Activity log
        log_entry = {
            "type": "PANEL_RECORD_CREATED",
            "panelId": ObjectId(panel_id),
            "panelName": panel.get("name", ""),
            "recordId": result.inserted_id,
            "sellerId": ObjectId(seller_id),
            "createdBy": user_id,
            "timestamp": now,
        }
        # Include product reference if panel is linked to inventory
        relation_fields = [f for f in panel.get("fields", []) if f["type"] == "relation" and f.get("relatedPanel") == "inventory"]
        for rf in relation_fields:
            if clean_data.get(rf["key"]):
                log_entry["productId"] = clean_data[rf["key"]]
                break
        # Include unique fields in log (e.g. QC number)
        for f in panel.get("fields", []):
            if f.get("unique") and clean_data.get(f["key"]):
                log_entry[f["key"]] = clean_data[f["key"]]
        await db.panel_activity_logs.insert_one(log_entry)

        logger.info(f"Record created in panel {panel_id} by {user_id}")

        # Execute automation rules (custom panels only)
        if automation_executor:
            try:
                await automation_executor(clean_data, panel_id, str(result.inserted_id), seller_id, user_id, "record_created")
            except Exception as e:
                logger.error(f"Automation execution error: {e}")

        return serialize_doc(doc)

    # ── UPDATE RECORD ──
    @router.put("/panels/{panel_id}/records/{record_id}")
    async def update_record(panel_id: str, record_id: str, data: UpdateRecordRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "edit")
        seller_id = await get_seller_id(user)

        panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        record = await db.panel_records.find_one({"_id": ObjectId(record_id), "panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")

        await validate_record_data(panel, data.data, seller_id, exclude_record_id=record_id)

        field_keys = {f["key"] for f in panel.get("fields", []) if not f.get("disabled")}
        clean_data = {k: v for k, v in data.data.items() if k in field_keys}

        await db.panel_records.update_one(
            {"_id": ObjectId(record_id)},
            {"$set": {"data": clean_data, "updatedAt": datetime.now(timezone.utc)}}
        )
        logger.info(f"Record {record_id} updated in panel {panel_id}")

        # Execute automation rules on update too
        if automation_executor:
            user_id = str(user.get("_id") or user.get("id"))
            try:
                await automation_executor(clean_data, panel_id, record_id, seller_id, user_id, "record_updated")
            except Exception as e:
                logger.error(f"Automation execution error on update: {e}")

        return {"message": "Record updated"}

    # ── DELETE RECORD ──
    @router.delete("/panels/{panel_id}/records/{record_id}")
    async def delete_record(panel_id: str, record_id: str, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "edit")
        seller_id = await get_seller_id(user)

        record = await db.panel_records.find_one({"_id": ObjectId(record_id), "panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")

        # Check if any other panel record has a relation pointing to this record
        # Check all panels' relation fields that target this panel
        panels_with_relations = db.panels.find({
            "sellerId": ObjectId(seller_id),
            "fields.type": "relation",
            "fields.relatedPanel": panel_id
        })
        async for p in panels_with_relations:
            for f in p.get("fields", []):
                if f.get("type") == "relation" and f.get("relatedPanel") == panel_id:
                    linked = await db.panel_records.find_one({
                        "panelId": p["_id"],
                        "sellerId": ObjectId(seller_id),
                        f"data.{f['key']}": record_id
                    })
                    if linked:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Cannot delete: this record is linked from panel '{p['name']}'. Remove the link first."
                        )

        await db.panel_records.delete_one({"_id": ObjectId(record_id)})
        logger.info(f"Record {record_id} deleted from panel {panel_id}")
        return {"message": "Record deleted"}

    # ── RELATION LOOKUP: search linkable entities ──
    @router.get("/panels/{panel_id}/relation-lookup")
    async def relation_lookup(
        panel_id: str,
        target: str = "",
        search: str = "",
        authorization: str = Header(...)
    ):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        if not target:
            return {"results": []}

        results = []
        limit = 20

        if target == "inventory":
            # Inventory items are in sellerListings, joined with products for names
            pipeline = [
                {"$match": {"sellerId": ObjectId(seller_id), "status": {"$in": ["active", "paused"]}}},
                {"$lookup": {
                    "from": "products",
                    "localField": "productId",
                    "foreignField": "_id",
                    "as": "productData"
                }},
                {"$unwind": {"path": "$productData", "preserveNullAndEmptyArrays": True}},
                {"$project": {
                    "listingId": "$_id",
                    "productName": "$productData.name",
                    "sku": {"$ifNull": ["$sku", ""]},
                }},
            ]
            if search:
                pipeline.append({"$match": {
                    "$or": [
                        {"productName": {"$regex": search, "$options": "i"}},
                        {"sku": {"$regex": search, "$options": "i"}},
                    ]
                }})
            pipeline.append({"$limit": limit})
            items = await db.sellerListings.aggregate(pipeline).to_list(limit)
            for item in items:
                results.append({
                    "id": str(item.get("listingId", item.get("_id"))),
                    "label": item.get("productName", ""),
                    "sub": item.get("sku", ""),
                })

        elif target == "invoices":
            q = {"sellerId": ObjectId(seller_id)}
            if search:
                q["$or"] = [
                    {"invoiceNumber": {"$regex": search, "$options": "i"}},
                    {"buyerName": {"$regex": search, "$options": "i"}},
                ]
            cursor = db.invoices.find(q, {"_id": 1, "invoiceNumber": 1, "buyerName": 1}).sort("createdAt", -1).limit(limit)
            async for doc in cursor:
                results.append({"id": str(doc["_id"]), "label": doc.get("invoiceNumber", ""), "sub": doc.get("buyerName", "")})

        elif target == "buyers":
            q = {"sellerId": ObjectId(seller_id)}
            if search:
                q["$or"] = [
                    {"buyerName": {"$regex": search, "$options": "i"}},
                    {"phone": {"$regex": search, "$options": "i"}},
                    {"company": {"$regex": search, "$options": "i"}},
                ]
            cursor = db.seller_buyers.find(q, {"_id": 1, "buyerName": 1, "phone": 1}).sort("createdAt", -1).limit(limit)
            async for doc in cursor:
                results.append({"id": str(doc["_id"]), "label": doc.get("buyerName", ""), "sub": doc.get("phone", "")})

        elif target == "suppliers":
            q = {"sellerId": ObjectId(seller_id)}
            if search:
                q["$or"] = [
                    {"supplierName": {"$regex": search, "$options": "i"}},
                    {"phone": {"$regex": search, "$options": "i"}},
                    {"contact": {"$regex": search, "$options": "i"}},
                ]
            cursor = db.seller_suppliers.find(q, {"_id": 1, "supplierName": 1, "phone": 1}).sort("createdAt", -1).limit(limit)
            async for doc in cursor:
                results.append({"id": str(doc["_id"]), "label": doc.get("supplierName", ""), "sub": doc.get("phone", "")})

        elif target == "purchase_orders":
            q = {"sellerId": ObjectId(seller_id)}
            if search:
                q["$or"] = [
                    {"poNumber": {"$regex": search, "$options": "i"}},
                    {"supplierName": {"$regex": search, "$options": "i"}},
                ]
            cursor = db.purchase_orders.find(q, {"_id": 1, "poNumber": 1, "supplierName": 1}).sort("createdAt", -1).limit(limit)
            async for doc in cursor:
                results.append({"id": str(doc["_id"]), "label": doc.get("poNumber", ""), "sub": doc.get("supplierName", "")})

        elif target == "quotations":
            pipeline = [
                {"$match": {"sellerId": ObjectId(seller_id)}},
                {"$lookup": {"from": "seller_buyers", "localField": "buyerId", "foreignField": "_id", "as": "buyer"}},
                {"$unwind": {"path": "$buyer", "preserveNullAndEmptyArrays": True}},
                {"$project": {"quotationNumber": 1, "buyerName": "$buyer.buyerName"}},
            ]
            if search:
                pipeline.append({"$match": {"$or": [
                    {"quotationNumber": {"$regex": search, "$options": "i"}},
                    {"buyerName": {"$regex": search, "$options": "i"}},
                ]}})
            pipeline.extend([{"$sort": {"_id": -1}}, {"$limit": limit}])
            items = await db.quotations.aggregate(pipeline).to_list(limit)
            for item in items:
                results.append({"id": str(item["_id"]), "label": item.get("quotationNumber", ""), "sub": item.get("buyerName", "")})

        elif target == "composite_products":
            q = {"sellerId": ObjectId(seller_id)}
            if search:
                q["name"] = {"$regex": search, "$options": "i"}
            cursor = db.composite_products.find(q, {"_id": 1, "name": 1, "price": 1}).sort("createdAt", -1).limit(limit)
            async for doc in cursor:
                results.append({"id": str(doc["_id"]), "label": doc.get("name", ""), "sub": f"Rs {doc['price']}" if doc.get("price") else ""})

        elif target == "employees":
            q = {"companyId": ObjectId(seller_id), "employeeStatus": {"$in": ["active", "disabled"]}}
            if search:
                q["$or"] = [
                    {"name": {"$regex": search, "$options": "i"}},
                    {"email": {"$regex": search, "$options": "i"}},
                ]
            cursor = db.users.find(q, {"_id": 1, "name": 1, "email": 1, "employeeRole": 1}).limit(limit)
            async for doc in cursor:
                results.append({"id": str(doc["_id"]), "label": doc.get("name", doc.get("email", "")), "sub": doc.get("employeeRole", "")})

        else:
            # Custom panel records
            linked_panel = await db.panels.find_one({"_id": ObjectId(target), "sellerId": ObjectId(seller_id)})
            if linked_panel:
                q = {"panelId": ObjectId(target), "sellerId": ObjectId(seller_id)}
                if search:
                    text_keys = [f"data.{f['key']}" for f in linked_panel.get("fields", []) if f["type"] in ("text", "dropdown")]
                    if text_keys:
                        q["$or"] = [{k: {"$regex": search, "$options": "i"}} for k in text_keys]
                cursor = db.panel_records.find(q).sort("createdAt", -1).limit(limit)
                async for doc in cursor:
                    data = doc.get("data", {})
                    label = ""
                    for f in linked_panel.get("fields", []):
                        if f["type"] in ("text", "dropdown") and data.get(f["key"]):
                            label = str(data[f["key"]])
                            break
                    results.append({"id": str(doc["_id"]), "label": label or str(doc["_id"])[:8], "sub": ""})

        return {"results": results}

    # ── SUPER ADMIN: SET ACCESS LEVEL ──
    @router.put("/admin/set-access-level")
    async def set_access_level(authorization: str = Header(...), data: dict = {}):
        user = await get_current_user(authorization)
        if not is_platform_admin(user):
            raise HTTPException(status_code=403, detail="Platform admin only")

        user_id = data.get("userId")
        level = data.get("level")
        if not user_id or level not in ("none", "standard", "advanced"):
            raise HTTPException(status_code=400, detail="userId and level (none/standard/advanced) required")

        try:
            result = await db.users.update_one(
                {"_id": ObjectId(user_id)},
                {"$set": {"businessToolAccess": level, "updatedAt": datetime.now(timezone.utc)}}
            )
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid user ID")

        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")

        logger.info(f"Access level set to '{level}' for user {user_id}")
        return {"message": f"Access level set to '{level}'"}

    # ── GET ACCESS LEVEL ──
    @router.get("/access-level")
    async def get_access_level(authorization: str = Header(...)):
        user = await get_current_user(authorization)
        level = user.get("businessToolAccess", "standard")
        if is_platform_admin(user):
            level = "advanced"
        return {"level": level, "limits": {"maxPanels": MAX_PANELS_PER_BUSINESS, "maxFieldsPerPanel": MAX_FIELDS_PER_PANEL}}

    # ── PANEL ACTIVITY LOGS ──
    @router.get("/panels/{panel_id}/activity-logs")
    async def get_panel_activity_logs(panel_id: str, authorization: str = Header(...), limit: int = 50):
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        cursor = db.panel_activity_logs.find(
            {"panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)},
            {"_id": 0, "type": 1, "panelName": 1, "recordId": 1, "productId": 1,
             "createdBy": 1, "timestamp": 1}
        ).sort("timestamp", -1).limit(limit)
        logs = await cursor.to_list(limit)

        for log in logs:
            if "recordId" in log:
                log["recordId"] = str(log["recordId"])
        return {"logs": serialize_doc(logs)}

    # ═══════════════════════════════════════════
    # DOCUMENT BUILDER — Phase 3B (READ-ONLY)
    # ═══════════════════════════════════════════

    @router.get("/panels/{panel_id}/export/excel")
    async def export_excel(panel_id: str, authorization: str = Header(...)):
        """Export panel records to Excel. READ-ONLY — no automation triggered."""
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        try:
            panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid panel ID format")
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        records = await db.panel_records.find(
            {"panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)}
        ).sort("createdAt", -1).to_list(5000)

        fields = [f for f in panel.get("fields", []) if not f.get("disabled")]

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = panel["name"][:31]

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Write headers
        headers = ["#"] + [f["label"] for f in fields] + ["Created At"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Resolve relation fields
        for row_idx, rec in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            for col_idx, f in enumerate(fields, 2):
                val = rec.get("data", {}).get(f["key"], "")
                # Resolve relations
                if f["type"] == "relation" and val:
                    resolved = await resolve_relation_display(seller_id, f, val)
                    val = resolved.get("label", str(val)) if resolved else str(val)
                elif isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                elif isinstance(val, bool):
                    val = "Yes" if val else "No"
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
                cell.border = thin_border
            # Created At
            created = rec.get("createdAt", "")
            if isinstance(created, datetime):
                created = created.strftime("%Y-%m-%d %H:%M")
            ws.cell(row=row_idx, column=len(fields) + 2, value=str(created)).border = thin_border

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{panel['name'].replace(' ', '_')}_export.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    @router.get("/panels/{panel_id}/export/pdf")
    async def export_pdf(panel_id: str, authorization: str = Header(...)):
        """Export panel records to PDF. READ-ONLY — no automation triggered."""
        user = await get_current_user(authorization)
        check_panel_access(user, panel_id, "view")
        seller_id = await get_seller_id(user)

        try:
            panel = await db.panels.find_one({"_id": ObjectId(panel_id), "sellerId": ObjectId(seller_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid panel ID format")
        if not panel:
            raise HTTPException(status_code=404, detail="Panel not found")

        records = await db.panel_records.find(
            {"panelId": ObjectId(panel_id), "sellerId": ObjectId(seller_id)}
        ).sort("createdAt", -1).to_list(1000)

        fields = [f for f in panel.get("fields", []) if not f.get("disabled")]

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch

        output = io.BytesIO()
        page_size = landscape(A4) if len(fields) > 5 else A4
        doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"<b>{panel['name']}</b>", styles['Title']))
        if panel.get("description"):
            elements.append(Paragraph(panel["description"], styles['Normal']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Total Records: {len(records)} | Exported: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Table
        header_row = ["#"] + [f["label"] for f in fields]
        table_data = [header_row]

        for i, rec in enumerate(records, 1):
            row = [str(i)]
            for f in fields:
                val = rec.get("data", {}).get(f["key"], "")
                if f["type"] == "relation" and val:
                    resolved = await resolve_relation_display(seller_id, f, val)
                    val = resolved.get("label", str(val)) if resolved else str(val)
                elif isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                elif isinstance(val, bool):
                    val = "Yes" if val else "No"
                row.append(str(val) if val else "")
            table_data.append(row)

        if len(table_data) > 1:
            col_count = len(header_row)
            avail_width = page_size[0] - inch
            col_width = avail_width / col_count

            t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("No records found.", styles['Normal']))

        doc.build(elements)
        output.seek(0)

        filename = f"{panel['name'].replace(' ', '_')}_export.pdf"
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    return router
