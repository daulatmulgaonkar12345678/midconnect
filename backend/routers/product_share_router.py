"""
Product Share & WhatsApp Document Sharing Router
- Product catalog generation (PDF/Excel)
- Secure document links for WhatsApp sharing
- Recipient selection (buyers + suppliers)
"""

from fastapi import APIRouter, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from bson import ObjectId
from pydantic import BaseModel
import io
import csv
import os
import secrets
import urllib.parse
import logging

from utils.permissions import authenticate_user, resolve_seller_id

logger = logging.getLogger(__name__)


class ProductShareRequest(BaseModel):
    productIds: List[str]
    recipientIds: List[str]
    format: str = "pdf"  # pdf or xlsx
    showPrice: bool = True
    sendWhatsApp: bool = False
    recipientPhone: Optional[str] = None


class ShareDocumentRequest(BaseModel):
    documentType: str = "invoice"
    documentId: str = ""
    recipientPhone: str = ""
    message: str = ""


def serialize_doc(doc):
    if isinstance(doc, list):
        return [serialize_doc(d) for d in doc]
    if isinstance(doc, dict):
        result = {}
        for k, v in doc.items():
            if k == "_id":
                result["id"] = str(v)
            elif isinstance(v, ObjectId):
                result[k] = str(v)
            elif isinstance(v, datetime):
                result[k] = v.isoformat()
            else:
                result[k] = v
        return result
    return doc


def init_product_share_router(db, verify_token_func):
    router = APIRouter(tags=["Product Sharing"])

    async def get_current_user(authorization: str):
        return await authenticate_user(db, verify_token_func, authorization)

    async def get_seller_id(user: dict) -> str:
        return resolve_seller_id(user)

    # ─── Get all recipients (buyers + suppliers) ───
    @router.get("/recipients")
    async def list_recipients(authorization: str = Header(...), search: Optional[str] = None):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        buyer_q = {"sellerId": ObjectId(seller_id)}
        supplier_q = {"sellerId": ObjectId(seller_id)}
        if search:
            buyer_q["$or"] = [{"buyerName": {"$regex": search, "$options": "i"}}, {"company": {"$regex": search, "$options": "i"}}]
            supplier_q["$or"] = [{"supplierName": {"$regex": search, "$options": "i"}}, {"contact": {"$regex": search, "$options": "i"}}]

        buyers = await db.seller_buyers.find(buyer_q, {"_id": 1, "buyerName": 1, "company": 1, "phone": 1, "email": 1}).sort("buyerName", 1).to_list(200)
        suppliers = await db.seller_suppliers.find(supplier_q, {"_id": 1, "supplierName": 1, "contact": 1, "phone": 1, "email": 1}).sort("supplierName", 1).to_list(200)

        recipients = []
        for b in buyers:
            recipients.append({"id": str(b["_id"]), "name": b.get("buyerName", "Unknown"), "company": b.get("company", ""), "phone": b.get("phone", ""), "type": "Buyer"})
        for s in suppliers:
            recipients.append({"id": str(s["_id"]), "name": s.get("supplierName", "Unknown"), "company": s.get("contact", ""), "phone": s.get("phone", ""), "type": "Supplier"})

        return {"recipients": recipients}

    # ─── Get product categories for grouping ───
    @router.get("/product-categories")
    async def get_product_categories(authorization: str = Header(...)):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "status": {"$in": ["active", "paused"]}}},
            {"$lookup": {"from": "products", "localField": "productId", "foreignField": "_id", "as": "prod"}},
            {"$unwind": {"path": "$prod", "preserveNullAndEmptyArrays": True}},
            {"$group": {"_id": {"$ifNull": ["$prod.categoryName", "Uncategorized"]}, "count": {"$sum": 1}}},
            {"$sort": {"_id": 1}}
        ]
        cats = await db.sellerListings.aggregate(pipeline).to_list(100)
        return {"categories": [{"name": c["_id"], "count": c["count"]} for c in cats]}

    # ─── Create product share / generate catalog ───
    @router.post("/product-shares")
    async def create_product_share(data: ProductShareRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        if not data.productIds:
            raise HTTPException(status_code=400, detail="No products selected")
        if not data.recipientIds:
            raise HTTPException(status_code=400, detail="No recipients selected")

        # Get seller profile
        seller = await db.users.find_one({"_id": ObjectId(seller_id)})
        profile = seller.get("profile", {}) if seller else {}
        billing = seller.get("billingSettings", {}) if seller else {}

        # Get catalog sharing settings
        catalog_settings = seller.get("catalogSettings", {}) if seller else {}

        # Get products with details
        product_oids = [ObjectId(pid) for pid in data.productIds if ObjectId.is_valid(pid)]
        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "productId": {"$in": product_oids}, "status": {"$in": ["active", "paused"]}}},
            {"$lookup": {"from": "products", "localField": "productId", "foreignField": "_id", "as": "prod"}},
            {"$unwind": {"path": "$prod", "preserveNullAndEmptyArrays": True}},
            {"$project": {
                "productId": 1,
                "selling_price": {"$ifNull": ["$selling_price", {"$ifNull": ["$minPrice", 0]}]},
                "productName": "$prod.name",
                "categoryName": "$prod.categoryName",
                "description": "$prod.description",
                "unit": {"$ifNull": ["$prod.unit", "piece"]},
                "moq": {"$ifNull": ["$moq", 1]},
                "images": {"$slice": [{"$ifNull": ["$images", []]}, 1]},
                "coverImageUrl": "$prod.coverImageUrl",
                "searchableAttributes": {"$ifNull": ["$searchableAttributes", {}]},
                "attributeLabels": {"$ifNull": ["$attributeLabels", {}]},
                "hsn": "$prod.hsn",
                "pricingTiers": {"$ifNull": ["$pricingTiers", []]},
            }}
        ]
        products = await db.sellerListings.aggregate(pipeline).to_list(500)

        # Post-process: build specification string from searchableAttributes
        for p in products:
            attrs = p.get("searchableAttributes", {})
            labels = p.get("attributeLabels", {})
            if attrs:
                spec_parts = []
                for key, val in attrs.items():
                    label = labels.get(key, key.replace("_", " ").title())
                    spec_parts.append(f"{label}: {val}")
                p["specification"] = ", ".join(spec_parts)
            else:
                p["specification"] = ""

            # Resolve image: prefer listing images, fallback to product coverImageUrl
            imgs = p.get("images", [])
            cover = p.get("coverImageUrl")
            if not imgs and cover:
                p["images"] = [cover]

            # Resolve selling price from pricingTiers if not set
            if not p.get("selling_price") and p.get("pricingTiers"):
                tiers = p["pricingTiers"]
                if tiers and isinstance(tiers, list) and len(tiers) > 0:
                    p["selling_price"] = tiers[0].get("pricePerUnit", 0)

        if not products:
            raise HTTPException(status_code=400, detail="No valid products found")

        # Generate catalog bytes
        now = datetime.now(timezone.utc)
        seller_info = {
            "businessName": profile.get("businessName", seller.get("email", "Seller")),
            "phone": profile.get("phone", ""),
            "email": seller.get("email", ""),
            "address": profile.get("address", ""),
            "city": profile.get("city", ""),
            "state": profile.get("state", ""),
            "gstNumber": profile.get("gstNumber", ""),
            "logoUrl": billing.get("companyLogoUrl", "") or profile.get("sellerLogoUrl", ""),
        }

        if data.format == "xlsx":
            catalog_bytes = generate_excel_catalog(products, seller_info, data.showPrice, catalog_settings)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"
        else:
            catalog_bytes = generate_pdf_catalog(products, seller_info, data.showPrice, catalog_settings)
            content_type = "application/pdf"
            ext = "pdf"

        # Create share record
        share_doc = {
            "sellerId": ObjectId(seller_id),
            "recipientIds": [ObjectId(rid) for rid in data.recipientIds if ObjectId.is_valid(rid)],
            "productIds": product_oids,
            "productCount": len(products),
            "format": data.format,
            "showPrice": data.showPrice,
            "catalogData": catalog_bytes,
            "contentType": content_type,
            "ext": ext,
            "createdAt": now,
        }
        result = await db.product_shares.insert_one(share_doc)
        share_id = str(result.inserted_id)

        # Generate secure document link
        token = secrets.token_urlsafe(32)
        await db.document_shares.insert_one({
            "token": token,
            "sellerId": ObjectId(seller_id),
            "documentType": "catalog",
            "documentId": share_id,
            "recipientPhone": data.recipientPhone or "",
            "expiresAt": now + timedelta(days=7),
            "createdAt": now,
        })

        # Build WhatsApp link if requested
        whatsapp_link = None
        if data.sendWhatsApp and data.recipientPhone:
            app_url = os.environ.get("FRONTEND_URL", "https://low-stock-admin.preview.emergentagent.com")
            doc_url = f"{app_url}/api/doc/{token}"
            msg = f"Hello,\n\nPlease find our product catalog below.\n\nDownload here:\n{doc_url}\n\nRegards,\n{seller_info['businessName']}"
            phone = data.recipientPhone.replace("+", "").replace(" ", "").replace("-", "")
            if not phone.startswith("91"):
                phone = "91" + phone
            whatsapp_link = f"https://wa.me/{phone}?text={urllib.parse.quote(msg)}"

        return {
            "shareId": share_id,
            "token": token,
            "downloadUrl": f"/api/business-tools/product-shares/{share_id}/download",
            "documentLink": f"/api/doc/{token}",
            "whatsappLink": whatsapp_link,
            "productCount": len(products),
            "recipientCount": len(data.recipientIds),
        }

    # ─── Download catalog by share ID (auth required) ───
    @router.get("/product-shares/{share_id}/download")
    async def download_catalog(share_id: str, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        share = await db.product_shares.find_one({"_id": ObjectId(share_id), "sellerId": ObjectId(seller_id)})
        if not share:
            raise HTTPException(status_code=404, detail="Share not found")

        filename = f"product-catalog-{datetime.now().strftime('%Y%m%d')}.{share.get('ext', 'pdf')}"
        return StreamingResponse(
            io.BytesIO(share["catalogData"]),
            media_type=share.get("contentType", "application/pdf"),
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ─── Generate secure link for any document (invoice, PO, catalog) ───
    @router.post("/share-document")
    async def share_document(data: ShareDocumentRequest, authorization: str = Header(...)):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)

        profile = (await db.users.find_one({"_id": ObjectId(seller_id)})) or {}
        biz_name = profile.get("profile", {}).get("businessName", "Seller")

        token = secrets.token_urlsafe(32)
        await db.document_shares.insert_one({
            "token": token,
            "sellerId": ObjectId(seller_id),
            "documentType": data.documentType,
            "documentId": data.documentId,
            "recipientPhone": data.recipientPhone,
            "expiresAt": now + timedelta(days=7),
            "createdAt": now,
        })

        app_url = os.environ.get("FRONTEND_URL", "https://low-stock-admin.preview.emergentagent.com")
        doc_url = f"{app_url}/api/doc/{token}"

        templates = {
            "invoice": f"Hello,\n\nYour invoice has been generated.\n\nDownload here:\n{doc_url}\n\nRegards,\n{biz_name}",
            "po": f"Hello,\n\nPlease find the purchase order.\n\nDownload here:\n{doc_url}\n\nRegards,\n{biz_name}",
            "catalog": f"Hello,\n\nPlease find our product catalog.\n\nDownload here:\n{doc_url}\n\nRegards,\n{biz_name}",
        }
        msg = data.message or templates.get(data.documentType, templates["catalog"])

        phone = data.recipientPhone.replace("+", "").replace(" ", "").replace("-", "")
        if phone and not phone.startswith("91"):
            phone = "91" + phone

        whatsapp_link = f"https://wa.me/{phone}?text={urllib.parse.quote(msg)}" if phone else None

        return {"token": token, "documentLink": f"/api/doc/{token}", "whatsappLink": whatsapp_link}

    # ─── Catalog sharing settings ───
    @router.get("/catalog-settings")
    async def get_catalog_settings(authorization: str = Header(...)):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        seller = await db.users.find_one({"_id": ObjectId(seller_id)})
        settings = (seller or {}).get("catalogSettings", {})
        defaults = {"showImage": True, "showName": True, "showCategory": True, "showSpecification": True, "showDescription": True, "showPrice": True, "showUnit": True, "showMoq": True}
        return {**defaults, **settings}

    @router.put("/catalog-settings")
    async def update_catalog_settings(authorization: str = Header(...), settings: dict = {}):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        allowed = ["showImage", "showName", "showCategory", "showSpecification", "showDescription", "showPrice", "showUnit", "showMoq"]
        update = {}
        for key in allowed:
            if key in settings:
                update[f"catalogSettings.{key}"] = bool(settings[key])
        if update:
            await db.users.update_one({"_id": ObjectId(seller_id)}, {"$set": update})
        return {"message": "Settings updated"}

    return router


# ─── PUBLIC ROUTE: Secure document download (no auth) ───
def init_public_doc_router(db):
    router = APIRouter(tags=["Public Documents"])

    @router.get("/doc/{token}")
    async def get_shared_document(token: str):
        try:
            share = await db.document_shares.find_one({"token": token})
        except Exception as e:
            logger.error(f"Document lookup error: {e}")
            raise HTTPException(status_code=500, detail="Internal error")

        if not share:
            raise HTTPException(status_code=404, detail="Document not found or link has expired")

        # Compare expiry safely (handle both naive and aware datetimes)
        expires_at = share.get("expiresAt")
        if expires_at:
            now_utc = datetime.now(timezone.utc)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at < now_utc:
                raise HTTPException(status_code=403, detail="Document link expired")

        doc_type = share.get("documentType", "catalog")
        doc_id = share.get("documentId", "")
        seller_id = share.get("sellerId")

        if not doc_id:
            raise HTTPException(status_code=404, detail="Document reference missing")

        try:
            if doc_type == "catalog":
                catalog = await db.product_shares.find_one({"_id": ObjectId(doc_id), "sellerId": seller_id})
                if not catalog or "catalogData" not in catalog:
                    raise HTTPException(status_code=404, detail="Catalog not found")
                filename = f"product-catalog.{catalog.get('ext', 'pdf')}"
                return StreamingResponse(
                    io.BytesIO(catalog["catalogData"]),
                    media_type=catalog.get("contentType", "application/pdf"),
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'}
                )

            elif doc_type == "invoice":
                from services.invoice_pdf_service import generate_invoice_pdf
                inv = await db.invoices.find_one({"_id": ObjectId(doc_id), "sellerId": seller_id})
                if not inv:
                    raise HTTPException(status_code=404, detail="Invoice not found")
                seller = await db.users.find_one({"_id": seller_id})
                profile = (seller or {}).get("profile", {})
                billing = (seller or {}).get("billingSettings", {})
                gst = (seller or {}).get("gst", {})
                bank_details = {k: billing.get(k, "") for k in ["bankName", "accountNumber", "accountName", "ifscCode", "branch", "upiId"]}
                seller_data = {
                    "businessName": profile.get("businessName", ""), "name": profile.get("businessName", ""),
                    "address": profile.get("address", ""), "city": profile.get("city", ""),
                    "state": profile.get("state", ""), "phone": profile.get("phone", ""),
                    "email": (seller or {}).get("email", ""), "gstNumber": gst.get("number", ""),
                    "sellerLogoUrl": billing.get("companyLogoUrl", "") or profile.get("sellerLogoUrl", ""),
                    "bankDetails": bank_details, "invoiceTerms": billing.get("invoiceTerms", ""),
                    "invoiceBackgroundImage": billing.get("invoiceBackgroundImage", ""),
                }
                buyer = await db.seller_buyers.find_one({"_id": inv.get("buyerId")}) or {}
                inv_ser = serialize_doc(inv)
                pdf_bytes = generate_invoice_pdf(inv_ser, seller_data, buyer, copy_type="original")
                filename = f"invoice-{inv.get('invoiceNumber', 'doc')}.pdf"
                return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

            elif doc_type == "po":
                po = await db.purchase_orders.find_one({"_id": ObjectId(doc_id), "sellerId": seller_id})
                if not po:
                    raise HTTPException(status_code=404, detail="Purchase order not found")
                po_data = serialize_doc(po)
                for k in ["_id", "sellerId"]:
                    po_data.pop(k, None)
                return po_data

            raise HTTPException(status_code=400, detail="Unsupported document type")

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Document retrieval error for token={token}: {e}")
            raise HTTPException(status_code=500, detail="Failed to retrieve document")

    return router


# ─── Catalog Generators ───

def generate_pdf_catalog(products: list, seller: dict, show_price: bool, settings: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import requests as req_lib

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("CatalogTitle", parent=styles["Title"], fontSize=18, spaceAfter=6, textColor=colors.HexColor("#1e3a5f"))
    subtitle_style = ParagraphStyle("CatalogSub", parent=styles["Normal"], fontSize=10, textColor=colors.gray, spaceAfter=12)
    header_style = ParagraphStyle("ColHeader", parent=styles["Normal"], fontSize=9, textColor=colors.white, fontName="Helvetica-Bold")
    cell_style = ParagraphStyle("CellText", parent=styles["Normal"], fontSize=8, leading=10)
    bold_cell = ParagraphStyle("BoldCell", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", leading=11)

    # Header
    elements.append(Paragraph("Product Catalog", title_style))
    elements.append(Paragraph(f"{seller.get('businessName', '')}", ParagraphStyle("BizName", parent=styles["Normal"], fontSize=12, fontName="Helvetica-Bold", spaceAfter=2)))

    contact_parts = []
    if seller.get("phone"):
        contact_parts.append(f"Phone: {seller['phone']}")
    if seller.get("email"):
        contact_parts.append(f"Email: {seller['email']}")
    if seller.get("address"):
        addr = seller["address"]
        if seller.get("city"):
            addr += f", {seller['city']}"
        if seller.get("state"):
            addr += f", {seller['state']}"
        contact_parts.append(addr)
    if seller.get("gstNumber"):
        contact_parts.append(f"GSTIN: {seller['gstNumber']}")
    if contact_parts:
        elements.append(Paragraph(" | ".join(contact_parts), subtitle_style))

    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d %b %Y')}  |  Total Products: {len(products)}", ParagraphStyle("DateLine", parent=styles["Normal"], fontSize=9, textColor=colors.gray, spaceAfter=10)))

    show = {
        "showImage": settings.get("showImage", True),
        "showName": settings.get("showName", True),
        "showCategory": settings.get("showCategory", True),
        "showSpecification": settings.get("showSpecification", True),
        "showDescription": settings.get("showDescription", True),
        "showPrice": settings.get("showPrice", True) and show_price,
        "showUnit": settings.get("showUnit", True),
        "showMoq": settings.get("showMoq", True),
    }

    # Download images into memory
    image_cache = {}
    if show["showImage"]:
        for p in products:
            imgs = p.get("images", [])
            img_url = imgs[0] if imgs else None
            if img_url:
                try:
                    resp = req_lib.get(img_url, timeout=5)
                    if resp.status_code == 200:
                        img_buf = io.BytesIO(resp.content)
                        image_cache[str(p.get("productId", ""))] = img_buf
                except Exception:
                    pass

    # Build table headers
    headers_list = ["#"]
    if show["showImage"]:
        headers_list.append("Image")
    if show["showName"]:
        headers_list.append("Product Name")
    if show["showCategory"]:
        headers_list.append("Category")
    if show["showSpecification"]:
        headers_list.append("Specification")
    if show["showDescription"]:
        headers_list.append("Description")
    if show["showPrice"]:
        headers_list.append("Price")
    if show["showUnit"]:
        headers_list.append("Unit")
    if show["showMoq"]:
        headers_list.append("MOQ")

    table_data = [[Paragraph(h, header_style) for h in headers_list]]

    for idx, p in enumerate(products, 1):
        row = [str(idx)]
        if show["showImage"]:
            pid = str(p.get("productId", ""))
            if pid in image_cache:
                try:
                    image_cache[pid].seek(0)
                    img = RLImage(image_cache[pid], width=12*mm, height=12*mm)
                    row.append(img)
                except Exception:
                    row.append("")
            else:
                row.append("")
        if show["showName"]:
            row.append(Paragraph(p.get("productName", "N/A"), bold_cell))
        if show["showCategory"]:
            row.append(Paragraph(p.get("categoryName", "") or "", cell_style))
        if show["showSpecification"]:
            spec = p.get("specification", "") or ""
            row.append(Paragraph(spec[:120], cell_style))
        if show["showDescription"]:
            desc = p.get("description", "") or ""
            row.append(Paragraph(desc[:150], cell_style))
        if show["showPrice"]:
            price = p.get("selling_price")
            row.append(f"Rs.{price:,.2f}" if price else "")
        if show["showUnit"]:
            row.append(p.get("unit", "") or "")
        if show["showMoq"]:
            moq = p.get("moq")
            row.append(str(moq) if moq else "")
        table_data.append(row)

    # Column widths
    n_cols = len(headers_list)
    available = 180 * mm
    col_widths = [available / n_cols] * n_cols
    col_widths[0] = 8 * mm
    if show["showImage"] and n_cols > 1:
        col_widths[1] = 16 * mm

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(t)

    # Footer
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("This catalog is generated for your reference. Prices and availability are subject to change.", ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.gray)))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def generate_excel_catalog(products: list, seller: dict, show_price: bool, settings: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Catalog"

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=f"Product Catalog - {seller.get('businessName', '')}")
    title_cell.font = Font(bold=True, size=14, color="1e3a5f")

    ws.merge_cells("A2:H2")
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y')}  |  Products: {len(products)}")
    date_cell.font = Font(size=10, color="666666")

    show = {
        "showName": settings.get("showName", True),
        "showCategory": settings.get("showCategory", True),
        "showSpecification": settings.get("showSpecification", True),
        "showDescription": settings.get("showDescription", True),
        "showPrice": settings.get("showPrice", True) and show_price,
        "showUnit": settings.get("showUnit", True),
        "showMoq": settings.get("showMoq", True),
    }

    headers = ["#"]
    if show["showName"]:
        headers.append("Product Name")
    if show["showCategory"]:
        headers.append("Category")
    if settings.get("showImage", True):
        headers.append("Image URL")
    if show["showSpecification"]:
        headers.append("Specification")
    if show["showDescription"]:
        headers.append("Description")
    if show["showPrice"]:
        headers.append("Selling Price")
    if show["showUnit"]:
        headers.append("Unit")
    if show["showMoq"]:
        headers.append("MOQ")

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for idx, p in enumerate(products, 1):
        row = idx + 4
        col = 1
        ws.cell(row=row, column=col, value=idx).border = thin_border
        col += 1
        if show["showName"]:
            c = ws.cell(row=row, column=col, value=p.get("productName", ""))
            c.font = Font(bold=True)
            c.border = thin_border
            col += 1
        if show["showCategory"]:
            ws.cell(row=row, column=col, value=p.get("categoryName", "")).border = thin_border
            col += 1
        if settings.get("showImage", True):
            imgs = p.get("images", [])
            ws.cell(row=row, column=col, value=imgs[0] if imgs else "").border = thin_border
            col += 1
        if show["showSpecification"]:
            ws.cell(row=row, column=col, value=p.get("specification", "")).border = thin_border
            col += 1
        if show["showDescription"]:
            ws.cell(row=row, column=col, value=p.get("description", "")).border = thin_border
            col += 1
        if show["showPrice"]:
            price = p.get("selling_price")
            ws.cell(row=row, column=col, value=price if price else "").border = thin_border
            col += 1
        if show["showUnit"]:
            ws.cell(row=row, column=col, value=p.get("unit", "")).border = thin_border
            col += 1
        if show["showMoq"]:
            moq = p.get("moq")
            ws.cell(row=row, column=col, value=moq if moq else "").border = thin_border

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(5, len(products) + 5):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
