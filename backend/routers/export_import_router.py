"""
Export/Import Router - CSV, Excel, PDF exports for reports and bulk data import
"""

from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import io
import csv
import json
import logging

from utils.permissions import authenticate_user, resolve_seller_id

logger = logging.getLogger(__name__)


def init_export_import_router(db, verify_token_func):
    router = APIRouter(tags=["Export/Import"])

    async def get_current_user(authorization: str):
        return await authenticate_user(db, verify_token_func, authorization)

    async def get_seller_id(user: dict) -> str:
        return resolve_seller_id(user)

    def parse_date(date_str: Optional[str]) -> Optional[datetime]:
        if not date_str:
            return None
        try:
            return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        except Exception:
            return None

    REPORT_STATUSES = ["draft", "sent", "viewed", "partially_paid", "paid", "overdue"]

    def make_csv_response(rows: list, headers: list, filename: str):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    def make_excel_response(rows: list, headers: list, filename: str, sheet_name: str = "Report"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(rows) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ─── EXPORT ENDPOINTS ───

    @router.get("/export/sales")
    async def export_sales(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None,
        period: str = "monthly"
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        # Get detailed invoice data for export
        invoices = await db.invoices.find({
            "sellerId": ObjectId(seller_id),
            "createdAt": {"$gte": start, "$lt": end},
            "status": {"$in": REPORT_STATUSES}
        }).sort("createdAt", -1).to_list(5000)

        headers = ["Invoice No", "Date", "Buyer Name", "GSTIN", "Product", "HSN", "Qty", "Rate", "Taxable Amount", "CGST", "SGST", "IGST", "Total Amount", "Payment Status"]
        rows = []

        for inv in invoices:
            buyer = await db.seller_buyers.find_one({"_id": inv.get("buyerId")})
            buyer_name = buyer.get("buyerName", "Unknown") if buyer else "Unknown"
            buyer_gstin = buyer.get("gstNumber", "") if buyer else ""
            inv_date = inv.get("createdAt", "").strftime("%d/%m/%Y") if inv.get("createdAt") else ""

            for item in inv.get("items", []):
                rows.append([
                    inv.get("invoiceNumber", ""),
                    inv_date,
                    buyer_name,
                    buyer_gstin,
                    item.get("productName", ""),
                    item.get("hsn", ""),
                    item.get("quantity", 0),
                    round(item.get("price", 0), 2),
                    round(item.get("taxableAmount", item.get("total", 0)), 2),
                    round(item.get("cgst", 0), 2),
                    round(item.get("sgst", 0), 2),
                    round(item.get("igst", 0), 2),
                    round(item.get("total", 0), 2),
                    inv.get("status", "draft")
                ])

        filename = f"sales-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Sales Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/profit")
    async def export_profit(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "createdAt": {"$gte": start, "$lt": end}, "status": {"$in": REPORT_STATUSES}}},
            {"$unwind": "$items"},
            {"$group": {
                "_id": "$items.productName",
                "totalQuantity": {"$sum": "$items.quantity"},
                "totalRevenue": {"$sum": "$items.total"},
                "totalCost": {"$sum": {"$multiply": [{"$ifNull": ["$items.purchase_price", 0]}, "$items.quantity"]}},
                "invoiceCount": {"$sum": 1}
            }},
            {"$sort": {"totalRevenue": -1}}
        ]
        results = await db.invoices.aggregate(pipeline).to_list(500)

        headers = ["Product", "Qty Sold", "Revenue", "Cost", "Profit", "Margin %", "Invoice Count"]
        rows = []
        for r in results:
            rev = r.get("totalRevenue", 0)
            cost = r.get("totalCost", 0)
            profit = rev - cost
            margin = (profit / rev * 100) if rev > 0 else 0
            rows.append([r["_id"], r["totalQuantity"], round(rev, 2), round(cost, 2), round(profit, 2), round(margin, 1), r["invoiceCount"]])

        filename = f"profit-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Profit Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/inventory")
    async def export_inventory(
        authorization: str = Header(...),
        format: str = "csv"
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "status": {"$in": ["active", "paused"]}}},
            {"$lookup": {"from": "products", "localField": "productId", "foreignField": "_id", "as": "prod"}},
            {"$unwind": {"path": "$prod", "preserveNullAndEmptyArrays": True}},
            {"$project": {
                "productName": "$prod.name",
                "sku": 1, "stock": {"$ifNull": ["$stock", 0]},
                "lowStockAlert": {"$ifNull": ["$lowStockAlert", 10]},
                "purchase_price": {"$ifNull": ["$purchase_price", 0]},
                "selling_price": {"$ifNull": ["$selling_price", 0]},
                "stockValue": {"$multiply": [{"$ifNull": ["$purchase_price", 0]}, {"$ifNull": ["$stock", 0]}]},
                "potentialRevenue": {"$multiply": [{"$ifNull": ["$selling_price", 0]}, {"$ifNull": ["$stock", 0]}]}
            }},
            {"$sort": {"productName": 1}}
        ]
        items = await db.sellerListings.aggregate(pipeline).to_list(500)

        headers = ["Product Name", "SKU", "Stock", "Low Stock Alert", "Purchase Price", "Selling Price", "Stock Value", "Potential Revenue", "Status"]
        rows = []
        for i in items:
            stock = i.get("stock", 0)
            alert = i.get("lowStockAlert", 10)
            status = "Out of Stock" if stock == 0 else ("Low Stock" if stock <= alert else "In Stock")
            rows.append([
                i.get("productName", "N/A"), i.get("sku", ""), stock, alert,
                round(i.get("purchase_price", 0), 2), round(i.get("selling_price", 0), 2),
                round(i.get("stockValue", 0), 2), round(i.get("potentialRevenue", 0), 2), status
            ])

        filename = f"inventory-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Inventory Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/buyers")
    async def export_buyers(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "createdAt": {"$gte": start, "$lt": end}, "status": {"$in": REPORT_STATUSES}}},
            {"$group": {"_id": "$buyerId", "totalSpent": {"$sum": "$total"}, "invoiceCount": {"$sum": 1}, "lastDate": {"$max": "$createdAt"}}},
            {"$sort": {"totalSpent": -1}}
        ]
        results = await db.invoices.aggregate(pipeline).to_list(500)

        headers = ["Buyer Name", "Company", "GSTIN", "Phone", "Email", "Total Spent", "Invoice Count", "Last Order"]
        rows = []
        for r in results:
            buyer = await db.seller_buyers.find_one({"_id": r["_id"]})
            if not buyer:
                continue
            rows.append([
                buyer.get("buyerName", "Unknown"), buyer.get("company", ""),
                buyer.get("gstNumber", ""), buyer.get("phone", ""), buyer.get("email", ""),
                round(r["totalSpent"], 2), r["invoiceCount"],
                r["lastDate"].strftime("%d/%m/%Y") if r.get("lastDate") else ""
            ])

        filename = f"buyers-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Buyers Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/invoices")
    async def export_invoices(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        invoices = await db.invoices.find({
            "sellerId": ObjectId(seller_id),
            "createdAt": {"$gte": start, "$lt": end},
            "status": {"$in": REPORT_STATUSES}
        }).sort("createdAt", -1).to_list(5000)

        headers = ["Invoice No", "Date", "Buyer Name", "Buyer GSTIN", "Subtotal", "GST", "Total", "Status", "Due Date", "Items Count"]
        rows = []
        for inv in invoices:
            buyer = await db.seller_buyers.find_one({"_id": inv.get("buyerId")})
            rows.append([
                inv.get("invoiceNumber", ""),
                inv.get("createdAt", "").strftime("%d/%m/%Y") if inv.get("createdAt") else "",
                buyer.get("buyerName", "Unknown") if buyer else "Unknown",
                buyer.get("gstNumber", "") if buyer else "",
                round(inv.get("subtotal", 0), 2),
                round(inv.get("gst", 0), 2),
                round(inv.get("total", 0), 2),
                inv.get("status", "draft"),
                inv.get("dueDate", "").strftime("%d/%m/%Y") if inv.get("dueDate") else "",
                len(inv.get("items", []))
            ])

        filename = f"invoices-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Invoice List")
        return make_csv_response(rows, headers, f"{filename}.csv")

    # ─── NEW REPORT EXPORTS ───

    @router.get("/export/outstanding")
    async def export_outstanding(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)

        outstanding_statuses = ["sent", "viewed", "partially_paid", "overdue"]
        match_stage: dict = {
            "sellerId": ObjectId(seller_id),
            "status": {"$in": outstanding_statuses}
        }
        start = parse_date(startDate)
        end = parse_date(endDate)
        if start or end:
            date_filter = {}
            if start:
                date_filter["$gte"] = start
            if end:
                date_filter["$lt"] = end + timedelta(days=1)
            match_stage["date"] = date_filter

        invoices = await db.invoices.find(match_stage).sort("date", -1).to_list(5000)

        headers = ["Invoice No", "Buyer Name", "Company", "Invoice Date", "Due Date",
                    "Total Amount", "Paid Amount", "Pending Amount", "Days Overdue", "Status", "Aging Bucket"]
        rows = []
        for inv in invoices:
            buyer = await db.seller_buyers.find_one({"_id": inv.get("buyerId")})
            inv_date = inv.get("date") or inv.get("createdAt")
            due_date = inv_date + timedelta(days=30) if inv_date else None
            days_overdue = max(0, (now - due_date).days) if due_date else 0
            paid = inv.get("totalPaid", 0)
            pending = inv.get("pendingAmount", inv.get("total", 0))
            status_label = "Partial" if inv.get("status") == "partially_paid" else "Unpaid"

            if days_overdue <= 0:
                bucket = "Current"
            elif days_overdue <= 30:
                bucket = "0-30 days"
            elif days_overdue <= 60:
                bucket = "31-60 days"
            elif days_overdue <= 90:
                bucket = "61-90 days"
            else:
                bucket = "90+ days"

            rows.append([
                inv.get("invoiceNumber", ""),
                buyer.get("buyerName", "Unknown") if buyer else "Unknown",
                buyer.get("company", "") if buyer else "",
                inv_date.strftime("%d/%m/%Y") if inv_date else "",
                due_date.strftime("%d/%m/%Y") if due_date else "",
                round(inv.get("total", 0), 2),
                round(paid, 2),
                round(pending, 2),
                days_overdue,
                status_label,
                bucket
            ])

        filename = f"outstanding-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Outstanding Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/purchase-orders")
    async def export_purchase_orders(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=30))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        purchase_statuses = ["sent", "confirmed", "partially_received", "received"]
        pos = await db.purchase_orders.find({
            "sellerId": ObjectId(seller_id),
            "status": {"$in": purchase_statuses},
            "createdAt": {"$gte": start, "$lt": end}
        }).sort("createdAt", -1).to_list(5000)

        headers = ["PO Number", "Supplier Name", "Phone", "Status", "Total Amount", "Items", "Date"]
        rows = []
        for po in pos:
            item_names = ", ".join(it.get("productName", "") for it in (po.get("items") or []))
            rows.append([
                po.get("poNumber", ""),
                po.get("supplierName", ""),
                po.get("supplierPhone", ""),
                po.get("status", ""),
                round(po.get("totalAmount", 0), 2),
                item_names,
                po.get("createdAt", "").strftime("%d/%m/%Y") if po.get("createdAt") else ""
            ])

        filename = f"purchase-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Purchase Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/stock-movement")
    async def export_stock_movement(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=30))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        seller_oid = ObjectId(seller_id)
        pipeline = [
            {"$match": {"sellerId": seller_oid}},
            {"$facet": {
                "opening": [
                    {"$match": {"createdAt": {"$lt": start}}},
                    {"$sort": {"createdAt": -1}},
                    {"$group": {"_id": "$listingId", "productName": {"$first": "$productName"}, "openingStock": {"$first": "$newStock"}}}
                ],
                "firstInRange": [
                    {"$match": {"createdAt": {"$gte": start, "$lt": end}}},
                    {"$sort": {"createdAt": 1}},
                    {"$group": {"_id": "$listingId", "productName": {"$first": "$productName"}, "firstPreviousStock": {"$first": "$previousStock"}}}
                ],
                "movements": [
                    {"$match": {"createdAt": {"$gte": start, "$lt": end}}},
                    {"$group": {
                        "_id": "$listingId",
                        "productName": {"$first": "$productName"},
                        "inward": {"$sum": {"$cond": [{"$in": ["$changeType", ["purchase", "purchase_receipt"]]}, {"$abs": "$quantity"}, 0]}},
                        "outward": {"$sum": {"$cond": [{"$eq": ["$changeType", "sale"]}, {"$abs": "$quantity"}, 0]}},
                        "adjPos": {"$sum": {"$cond": [{"$and": [{"$in": ["$changeType", ["adjustment", "damage"]]}, {"$gt": ["$quantity", 0]}]}, "$quantity", 0]}},
                        "adjNeg": {"$sum": {"$cond": [{"$and": [{"$in": ["$changeType", ["adjustment", "damage"]]}, {"$lt": ["$quantity", 0]}]}, "$quantity", 0]}}
                    }}
                ]
            }}
        ]

        result = await db.inventory_logs.aggregate(pipeline).to_list(1)
        if not result:
            headers = ["Product Name", "Opening Stock", "Inward", "Outward", "Adjustment", "Closing Stock"]
            filename = f"stock-movement-{datetime.now().strftime('%Y%m%d')}"
            if format == "xlsx":
                return make_excel_response([], headers, f"{filename}.xlsx", "Stock Movement")
            return make_csv_response([], headers, f"{filename}.csv")

        data = result[0]
        opening_map = {str(o["_id"]): o for o in data.get("opening", [])}
        first_map = {str(f["_id"]): f for f in data.get("firstInRange", [])}
        movement_map = {str(m["_id"]): m for m in data.get("movements", [])}

        all_ids = set(opening_map.keys()) | set(first_map.keys()) | set(movement_map.keys())

        headers = ["Product Name", "Opening Stock", "Inward", "Outward", "Adjustment", "Closing Stock"]
        rows = []
        for pid in all_ids:
            o = opening_map.get(pid)
            f = first_map.get(pid)
            m = movement_map.get(pid, {})

            opening = o["openingStock"] if o else (f["firstPreviousStock"] if f else 0)
            name = m.get("productName") or (o.get("productName") if o else None) or (f.get("productName") if f else "Unknown")
            inward = m.get("inward", 0)
            outward = m.get("outward", 0)
            adj = m.get("adjPos", 0) + m.get("adjNeg", 0)
            closing = opening + inward - outward + adj

            rows.append([name, opening, inward, outward, adj, closing])

        rows.sort(key=lambda x: x[3], reverse=True)

        filename = f"stock-movement-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Stock Movement")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/buyer-ledger")
    async def export_buyer_ledger(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "createdAt": {"$gte": start, "$lt": end}, "status": {"$in": REPORT_STATUSES}}},
            {"$group": {
                "_id": "$buyerId",
                "totalSales": {"$sum": "$total"},
                "totalPaid": {"$sum": {"$ifNull": ["$totalPaid", 0]}},
                "totalPending": {"$sum": {"$ifNull": ["$pendingAmount", "$total"]}},
                "invoiceCount": {"$sum": 1},
                "lastInvoiceDate": {"$max": "$createdAt"}
            }},
            {"$sort": {"totalSales": -1}}
        ]
        results = await db.invoices.aggregate(pipeline).to_list(5000)

        headers = ["Buyer Name", "Company", "Total Sales", "Total Paid", "Pending Amount", "Invoices", "Last Invoice Date"]
        rows = []
        for r in results:
            buyer = await db.seller_buyers.find_one({"_id": r["_id"]}) if r.get("_id") else {}
            buyer = buyer or {}
            rows.append([
                buyer.get("buyerName", "Unknown"),
                buyer.get("company", ""),
                round(r["totalSales"], 2),
                round(r["totalPaid"], 2),
                round(r["totalPending"], 2),
                r["invoiceCount"],
                r["lastInvoiceDate"].strftime("%d/%m/%Y") if r.get("lastInvoiceDate") else ""
            ])

        filename = f"buyer-ledger-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Buyer Ledger")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/product-performance")
    async def export_product_performance(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "createdAt": {"$gte": start, "$lt": end}, "status": {"$in": REPORT_STATUSES}}},
            {"$unwind": "$items"},
            {"$group": {
                "_id": "$items.productName",
                "quantitySold": {"$sum": "$items.quantity"},
                "revenue": {"$sum": "$items.total"},
                "cost": {"$sum": {"$multiply": [{"$ifNull": ["$items.purchase_price", 0]}, "$items.quantity"]}}
            }},
            {"$addFields": {"profit": {"$subtract": ["$revenue", "$cost"]}}},
            {"$sort": {"revenue": -1}}
        ]
        results = await db.invoices.aggregate(pipeline).to_list(5000)

        headers = ["Product Name", "Quantity Sold", "Revenue", "Profit", "Profit %"]
        rows = []
        for r in results:
            rev = r.get("revenue", 0)
            profit = r.get("profit", 0)
            rows.append([
                r["_id"] or "Unknown",
                r["quantitySold"],
                round(rev, 2),
                round(profit, 2),
                f"{round((profit / rev * 100) if rev > 0 else 0, 1)}%"
            ])

        filename = f"product-performance-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Product Performance")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/category-report")
    async def export_category_report(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=365))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        # Same logic as the report endpoint but simplified for export
        pipeline = [
            {"$match": {"sellerId": ObjectId(seller_id), "createdAt": {"$gte": start, "$lt": end}, "status": {"$in": REPORT_STATUSES}}},
            {"$unwind": "$items"},
            {"$project": {
                "productId": "$items.productId", "quantity": "$items.quantity",
                "revenue": "$items.total",
                "cost": {"$multiply": [{"$ifNull": ["$items.purchase_price", 0]}, "$items.quantity"]}
            }}
        ]
        items = await db.invoices.aggregate(pipeline).to_list(10000)

        product_ids = set()
        for item in items:
            pid = item.get("productId")
            if pid and pid not in ("none", "None", None):
                try:
                    product_ids.add(ObjectId(pid))
                except Exception:
                    pass

        category_map = {}
        if product_ids:
            products = await db.products.find({"_id": {"$in": list(product_ids)}}, {"_id": 1, "categoryName": 1, "categoryId": 1}).to_list(len(product_ids))
            cat_ids_needed = set()
            for p in products:
                if p.get("categoryName"):
                    category_map[str(p["_id"])] = p["categoryName"]
                elif p.get("categoryId"):
                    cat_ids_needed.add(p["categoryId"])
                    category_map[str(p["_id"])] = str(p["categoryId"])
            if cat_ids_needed:
                cats = await db.categories.find({"_id": {"$in": list(cat_ids_needed)}}, {"_id": 1, "name": 1}).to_list(len(cat_ids_needed))
                cmap = {str(c["_id"]): c["name"] for c in cats}
                for pid, cv in category_map.items():
                    if cv in cmap:
                        category_map[pid] = cmap[cv]

        cat_data: dict = {}
        for item in items:
            pid = item.get("productId")
            cat_name = category_map.get(str(pid), "Uncategorized") if pid else "Uncategorized"
            if cat_name not in cat_data:
                cat_data[cat_name] = {"qty": 0, "revenue": 0, "cost": 0}
            cat_data[cat_name]["qty"] += item.get("quantity", 0)
            cat_data[cat_name]["revenue"] += item.get("revenue", 0)
            cat_data[cat_name]["cost"] += item.get("cost", 0)

        headers = ["Category Name", "Total Sales (Qty)", "Revenue", "Profit", "Profit %"]
        rows = []
        for name, d in sorted(cat_data.items(), key=lambda x: x[1]["revenue"], reverse=True):
            profit = d["revenue"] - d["cost"]
            rows.append([
                name, d["qty"], round(d["revenue"], 2), round(profit, 2),
                f"{round((profit / d['revenue'] * 100) if d['revenue'] > 0 else 0, 1)}%"
            ])

        filename = f"category-report-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Category Report")
        return make_csv_response(rows, headers, f"{filename}.csv")

    @router.get("/export/low-stock")
    async def export_low_stock(
        authorization: str = Header(...),
        format: str = "csv",
        startDate: Optional[str] = None,
        endDate: Optional[str] = None
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)
        now = datetime.now(timezone.utc)
        start = parse_date(startDate) or (now - timedelta(days=30))
        end = parse_date(endDate)
        end = (end + timedelta(days=1)) if end else (now + timedelta(days=1))

        days_in_range = max(1, (end - start).days)
        seller_oid = ObjectId(seller_id)

        listings = await db.sellerListings.aggregate([
            {"$match": {"sellerId": seller_oid, "status": "active"}},
            {"$lookup": {"from": "products", "localField": "productId", "foreignField": "_id", "as": "prod"}},
            {"$unwind": {"path": "$prod", "preserveNullAndEmptyArrays": True}},
            {"$project": {"productName": {"$ifNull": ["$prod.name", "Unknown"]}, "stock": {"$ifNull": ["$stock", 0]}, "lowStockAlert": {"$ifNull": ["$lowStockAlert", 10]}}}
        ]).to_list(500)
        listing_map = {str(ls["_id"]): ls for ls in listings}

        consumption = await db.inventory_logs.aggregate([
            {"$match": {"sellerId": seller_oid, "changeType": "sale", "createdAt": {"$gte": start, "$lt": end}}},
            {"$group": {"_id": "$listingId", "totalSold": {"$sum": {"$abs": "$quantity"}}}}
        ]).to_list(1000)
        cons_map = {str(c["_id"]): c.get("totalSold", 0) for c in consumption}

        headers = ["Product Name", "Min Stock", "Current Stock", "Avg Consumption/Day", "Total Sold", "Status"]
        rows = []
        for lid, ls in listing_map.items():
            sold = cons_map.get(lid, 0)
            avg = round(sold / days_in_range, 2)
            stock = ls.get("stock", 0)
            min_s = ls.get("lowStockAlert", 10)
            status = "Out of Stock" if stock == 0 else ("Low Stock" if stock <= min_s else "Healthy")
            rows.append([ls.get("productName", "Unknown"), min_s, stock, avg, sold, status])

        rows.sort(key=lambda x: (0 if x[5] == "Out of Stock" else (1 if x[5] == "Low Stock" else 2), x[2]))

        filename = f"low-stock-analytics-{datetime.now().strftime('%Y%m%d')}"
        if format == "xlsx":
            return make_excel_response(rows, headers, f"{filename}.xlsx", "Low Stock Analytics")
        return make_csv_response(rows, headers, f"{filename}.csv")

    # ─── IMPORT TEMPLATES ───

    IMPORT_TEMPLATES = {
        "products": {
            "headers": ["Product Name", "HSN Code", "Category", "Unit", "Selling Price", "Purchase Price", "Stock Quantity", "Low Stock Alert", "Description"],
            "sample": [["Safety Gloves", "6116", "Safety Equipment", "pair", "120", "70", "100", "20", "Industrial safety gloves"],
                       ["Helmet", "6506", "Safety Equipment", "piece", "450", "280", "25", "10", "Hard hat helmet"]]
        },
        "inventory": {
            "headers": ["Product Name", "Stock Quantity", "Purchase Price", "Selling Price", "Low Stock Alert"],
            "sample": [["Safety Gloves", "100", "70", "120", "20"], ["Helmet", "25", "280", "450", "10"]]
        },
        "suppliers": {
            "headers": ["Supplier Name", "Company", "Phone", "Email", "GSTIN", "Address", "City", "State"],
            "sample": [["Raj Kumar", "Kumar Suppliers", "9876543210", "raj@kumar.com", "22AAAAA0000A1Z5", "MG Road", "Mumbai", "Maharashtra"]]
        },
        "buyers": {
            "headers": ["Buyer Name", "Company", "Phone", "Email", "GSTIN", "Address", "City", "State"],
            "sample": [["Amit Shah", "Shah Enterprises", "9876543210", "amit@shah.com", "22BBBBB0000B1Z5", "Park Street", "Kolkata", "West Bengal"]]
        }
    }

    @router.get("/import/template/{data_type}")
    async def download_import_template(
        data_type: str,
        authorization: str = Header(...),
        format: str = "csv"
    ):
        await get_current_user(authorization)
        template = IMPORT_TEMPLATES.get(data_type)
        if not template:
            raise HTTPException(status_code=400, detail=f"Invalid data type: {data_type}. Valid: {', '.join(IMPORT_TEMPLATES.keys())}")

        filename = f"{data_type}-import-template"
        if format == "xlsx":
            return make_excel_response(template["sample"], template["headers"], f"{filename}.xlsx", f"{data_type.title()} Template")
        return make_csv_response(template["sample"], template["headers"], f"{filename}.csv")

    @router.post("/import/validate")
    async def validate_import(
        authorization: str = Header(...),
        file: UploadFile = File(...),
        data_type: str = Form(...)
    ):
        await get_current_user(authorization)

        template = IMPORT_TEMPLATES.get(data_type)
        if not template:
            raise HTTPException(status_code=400, detail=f"Invalid data type: {data_type}")

        content = await file.read()
        rows = []
        errors = []

        try:
            if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if len(all_rows) < 2:
                    return {"valid": False, "errors": [{"row": 0, "message": "File has no data rows"}], "preview": [], "totalRows": 0}
                file_headers = [str(h).strip() for h in all_rows[0] if h is not None]
                data_rows = all_rows[1:]
            else:
                decoded = content.decode("utf-8-sig")
                reader = csv.reader(io.StringIO(decoded))
                all_rows = list(reader)
                if len(all_rows) < 2:
                    return {"valid": False, "errors": [{"row": 0, "message": "File has no data rows"}], "preview": [], "totalRows": 0}
                file_headers = [h.strip() for h in all_rows[0]]
                data_rows = all_rows[1:]
        except Exception as e:
            return {"valid": False, "errors": [{"row": 0, "message": f"Failed to parse file: {str(e)}"}], "preview": [], "totalRows": 0}

        # Validate headers
        expected = [h.lower() for h in template["headers"]]
        actual = [h.lower() for h in file_headers]
        missing = [h for h in expected if h not in actual]
        if missing:
            errors.append({"row": 0, "message": f"Missing columns: {', '.join(missing)}"})

        # Validate rows
        for idx, row in enumerate(data_rows):
            row_data = {}
            clean_row = [str(c).strip() if c is not None else "" for c in row]
            for col_idx, header in enumerate(file_headers):
                if col_idx < len(clean_row):
                    row_data[header.strip()] = clean_row[col_idx]

            row_num = idx + 2  # 1-indexed, accounting for header

            if data_type == "products":
                name = row_data.get("Product Name", "").strip()
                if not name:
                    errors.append({"row": row_num, "message": "Product Name is required"})
                price = row_data.get("Selling Price", "0")
                try:
                    float(price)
                except ValueError:
                    errors.append({"row": row_num, "message": f"Invalid Selling Price: {price}"})

            elif data_type == "inventory":
                name = row_data.get("Product Name", "").strip()
                if not name:
                    errors.append({"row": row_num, "message": "Product Name is required"})
                stock = row_data.get("Stock Quantity", "0")
                try:
                    int(float(stock))
                except ValueError:
                    errors.append({"row": row_num, "message": f"Invalid Stock Quantity: {stock}"})

            elif data_type in ("suppliers", "buyers"):
                name_key = "Supplier Name" if data_type == "suppliers" else "Buyer Name"
                name = row_data.get(name_key, "").strip()
                if not name:
                    errors.append({"row": row_num, "message": f"{name_key} is required"})
                gstin = row_data.get("GSTIN", "").strip()
                if gstin and len(gstin) != 15:
                    errors.append({"row": row_num, "message": f"Invalid GSTIN length: {gstin} (expected 15 chars)"})

            rows.append(row_data)

        # Check duplicates for products
        if data_type == "products":
            names = [r.get("Product Name", "").strip().lower() for r in rows if r.get("Product Name", "").strip()]
            seen = set()
            for i, n in enumerate(names):
                if n in seen:
                    errors.append({"row": i + 2, "message": f"Duplicate product: {n}"})
                seen.add(n)

        preview = rows[:20]  # Show first 20 rows

        return {
            "valid": len(errors) == 0,
            "errors": errors[:50],
            "preview": preview,
            "totalRows": len(rows),
            "headers": file_headers
        }

    @router.post("/import/process")
    async def process_import(
        authorization: str = Header(...),
        file: UploadFile = File(...),
        data_type: str = Form(...)
    ):
        user = await get_current_user(authorization)
        seller_id = await get_seller_id(user)

        template = IMPORT_TEMPLATES.get(data_type)
        if not template:
            raise HTTPException(status_code=400, detail=f"Invalid data type: {data_type}")

        content = await file.read()
        try:
            if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                file_headers = [str(h).strip() for h in all_rows[0] if h is not None]
                data_rows = all_rows[1:]
            else:
                decoded = content.decode("utf-8-sig")
                reader = csv.reader(io.StringIO(decoded))
                all_rows = list(reader)
                file_headers = [h.strip() for h in all_rows[0]]
                data_rows = all_rows[1:]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

        imported = 0
        skipped = 0
        errors_list = []
        now = datetime.now(timezone.utc)

        for idx, row in enumerate(data_rows):
            row_data = {}
            clean_row = [str(c).strip() if c is not None else "" for c in row]
            for col_idx, header in enumerate(file_headers):
                if col_idx < len(clean_row):
                    row_data[header.strip()] = clean_row[col_idx]

            try:
                if data_type == "products":
                    name = row_data.get("Product Name", "").strip()
                    if not name:
                        skipped += 1
                        continue

                    # Check if product exists
                    existing = await db.products.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "sellerId": ObjectId(seller_id)})
                    if existing:
                        skipped += 1
                        errors_list.append(f"Row {idx+2}: Product '{name}' already exists, skipped")
                        continue

                    product_doc = {
                        "name": name,
                        "sellerId": ObjectId(seller_id),
                        "hsn": row_data.get("HSN Code", ""),
                        "category": row_data.get("Category", ""),
                        "unit": row_data.get("Unit", "piece"),
                        "description": row_data.get("Description", ""),
                        "status": "active",
                        "createdAt": now,
                        "updatedAt": now
                    }
                    result = await db.products.insert_one(product_doc)

                    # Create seller listing
                    listing_doc = {
                        "productId": result.inserted_id,
                        "sellerId": ObjectId(seller_id),
                        "selling_price": float(row_data.get("Selling Price", 0) or 0),
                        "purchase_price": float(row_data.get("Purchase Price", 0) or 0),
                        "stock": int(float(row_data.get("Stock Quantity", 0) or 0)),
                        "lowStockAlert": int(float(row_data.get("Low Stock Alert", 10) or 10)),
                        "status": "active",
                        "createdAt": now,
                        "updatedAt": now
                    }
                    await db.sellerListings.insert_one(listing_doc)
                    imported += 1

                elif data_type == "inventory":
                    name = row_data.get("Product Name", "").strip()
                    if not name:
                        skipped += 1
                        continue

                    product = await db.products.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "sellerId": ObjectId(seller_id)})
                    if not product:
                        skipped += 1
                        errors_list.append(f"Row {idx+2}: Product '{name}' not found, skipped")
                        continue

                    update_fields = {"updatedAt": now}
                    stock = row_data.get("Stock Quantity", "")
                    if stock:
                        update_fields["stock"] = int(float(stock))
                    pp = row_data.get("Purchase Price", "")
                    if pp:
                        update_fields["purchase_price"] = float(pp)
                    sp = row_data.get("Selling Price", "")
                    if sp:
                        update_fields["selling_price"] = float(sp)
                    alert = row_data.get("Low Stock Alert", "")
                    if alert:
                        update_fields["lowStockAlert"] = int(float(alert))

                    await db.sellerListings.update_one(
                        {"productId": product["_id"], "sellerId": ObjectId(seller_id)},
                        {"$set": update_fields}
                    )
                    imported += 1

                elif data_type == "suppliers":
                    name = row_data.get("Supplier Name", "").strip()
                    if not name:
                        skipped += 1
                        continue

                    existing = await db.seller_suppliers.find_one({"supplierName": {"$regex": f"^{name}$", "$options": "i"}, "sellerId": ObjectId(seller_id)})
                    if existing:
                        skipped += 1
                        errors_list.append(f"Row {idx+2}: Supplier '{name}' already exists, skipped")
                        continue

                    doc = {
                        "sellerId": ObjectId(seller_id),
                        "supplierName": name,
                        "company": row_data.get("Company", ""),
                        "phone": row_data.get("Phone", ""),
                        "email": row_data.get("Email", ""),
                        "gstNumber": row_data.get("GSTIN", ""),
                        "address": row_data.get("Address", ""),
                        "city": row_data.get("City", ""),
                        "state": row_data.get("State", ""),
                        "status": "active",
                        "createdAt": now, "updatedAt": now
                    }
                    await db.seller_suppliers.insert_one(doc)
                    imported += 1

                elif data_type == "buyers":
                    name = row_data.get("Buyer Name", "").strip()
                    if not name:
                        skipped += 1
                        continue

                    existing = await db.seller_buyers.find_one({"buyerName": {"$regex": f"^{name}$", "$options": "i"}, "sellerId": ObjectId(seller_id)})
                    if existing:
                        skipped += 1
                        errors_list.append(f"Row {idx+2}: Buyer '{name}' already exists, skipped")
                        continue

                    doc = {
                        "sellerId": ObjectId(seller_id),
                        "buyerName": name,
                        "company": row_data.get("Company", ""),
                        "phone": row_data.get("Phone", ""),
                        "email": row_data.get("Email", ""),
                        "gstNumber": row_data.get("GSTIN", ""),
                        "address": row_data.get("Address", ""),
                        "city": row_data.get("City", ""),
                        "state": row_data.get("State", ""),
                        "status": "active",
                        "createdAt": now, "updatedAt": now
                    }
                    await db.seller_buyers.insert_one(doc)
                    imported += 1

            except Exception as e:
                skipped += 1
                errors_list.append(f"Row {idx+2}: {str(e)}")

        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(data_rows),
            "errors": errors_list[:50]
        }

    return router
